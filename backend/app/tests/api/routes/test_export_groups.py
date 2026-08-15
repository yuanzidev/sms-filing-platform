"""Tests for export groups API."""

import io
import uuid
from collections.abc import Generator

import pytest
from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook
from sqlmodel import Session, delete, select

from app.core.config import settings
from app.core.db import engine
from app.models import FilingTask, User


@pytest.fixture(scope="module", autouse=True)
def _cleanup_filing_tasks() -> Generator[None, None, None]:
    # 清理本模块创建的 filing_task 行，避免 conftest 清理 User 时触发外键约束
    yield
    with Session(engine) as session:
        session.execute(delete(FilingTask))
        session.commit()


def test_registry_returns_all_fields(
    client: TestClient, superuser_token_headers: dict[str, str]
) -> None:
    r = client.get(
        f"{settings.API_V1_STR}/export-groups/registry",
        headers=superuser_token_headers,
    )
    assert r.status_code == 200
    data = r.json()
    assert isinstance(data, list)
    assert len(data) > 30
    names = {item["name"] for item in data}
    assert "sms_signature" in names
    assert "signature_type" in names
    assert "cert_image" in names

    sig_field = next(item for item in data if item["name"] == "signature_type")
    assert sig_field["label"] == "签名类型/来源"
    assert sig_field["source"] == "qualification"
    assert sig_field["group"] == "签名与模板"


def test_registry_template_download(
    client: TestClient, superuser_token_headers: dict[str, str]
) -> None:
    r = client.get(
        f"{settings.API_V1_STR}/export-groups/registry/template",
        headers=superuser_token_headers,
    )
    assert r.status_code == 200
    assert "spreadsheetml" in r.headers["content-type"]
    wb = load_workbook(io.BytesIO(r.content))
    ws = wb.active
    assert ws.cell(row=1, column=1).value == "字段编码"
    assert ws.cell(row=1, column=2).value == "字段名称"
    assert ws.cell(row=1, column=3).value == "所属分组"
    codes = [ws.cell(row=i, column=1).value for i in range(2, ws.max_row + 1)]
    assert "sms_signature" in codes


def _create_group(
    client: TestClient, headers: dict[str, str], name: str = "测试导出组"
) -> str:
    r = client.post(
        f"{settings.API_V1_STR}/export-groups",
        headers=headers,
        json={
            "name": name,
            "description": "测试用",
            "fields": [
                {
                    "field_name": "sms_signature",
                    "field_label": "短信签名",
                    "sort_order": 0,
                },
                {
                    "field_name": "main_port_number",
                    "field_label": "主端口号",
                    "sort_order": 1,
                },
            ],
        },
    )
    assert r.status_code == 200
    return r.json()["id"]


def test_export_group_download(
    client: TestClient, superuser_token_headers: dict[str, str]
) -> None:
    gid = _create_group(client, superuser_token_headers)
    r = client.get(
        f"{settings.API_V1_STR}/export-groups/{gid}/export",
        headers=superuser_token_headers,
    )
    assert r.status_code == 200
    assert "spreadsheetml" in r.headers["content-type"]
    wb = load_workbook(io.BytesIO(r.content))
    ws = wb.active
    assert ws.cell(row=1, column=1).value == "字段组名称"
    assert ws.cell(row=2, column=1).value == "测试导出组"
    assert ws.cell(row=2, column=2).value == "sms_signature"
    assert ws.cell(row=3, column=2).value == "main_port_number"
    assert ws.cell(row=2, column=5).value == "是"


def test_export_group_download_404(
    client: TestClient, superuser_token_headers: dict[str, str]
) -> None:
    r = client.get(
        f"{settings.API_V1_STR}/export-groups/00000000-0000-0000-0000-000000000000/export",
        headers=superuser_token_headers,
    )
    assert r.status_code == 404


def _build_import_xlsx(rows: list[list]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "字段组"
    headers = ["字段组名称", "字段编码", "字段名称", "字段顺序", "是否启用"]
    ws.append(headers)
    for row in rows:
        ws.append(row)
    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def test_import_export_group_success(
    client: TestClient, superuser_token_headers: dict[str, str]
) -> None:
    content = _build_import_xlsx(
        [
            ["导入字段组A", "sms_signature", "短信签名", 0, "是"],
            ["导入字段组A", "main_port_number", "主端口号", 1, "是"],
        ]
    )
    r = client.post(
        f"{settings.API_V1_STR}/export-groups/import",
        headers=superuser_token_headers,
        files={
            "file": (
                "group.xlsx",
                content,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert r.status_code == 200
    data = r.json()
    assert data["success_count"] == 1
    assert data["group_name"] == "导入字段组A"
    assert data["field_count"] == 2

    groups = client.get(
        f"{settings.API_V1_STR}/export-groups", headers=superuser_token_headers
    ).json()
    g = next(g for g in groups["data"] if g["name"] == "导入字段组A")
    assert len(g["fields"]) == 2


def test_import_export_group_one_column_field_codes(
    client: TestClient, superuser_token_headers: dict[str, str]
) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "字段编码对照表"
    ws.append(["sms_signature"])
    ws.append(["main_port_number"])
    output = io.BytesIO()
    wb.save(output)

    r = client.post(
        f"{settings.API_V1_STR}/export-groups/import",
        headers=superuser_token_headers,
        files={
            "file": (
                "字段编码对照表.xlsx",
                output.getvalue(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert r.status_code == 200
    data = r.json()
    assert data["success_count"] == 1
    assert data["group_name"] == "字段编码对照表"
    assert data["field_count"] == 2


def test_import_export_group_registry_table_format(
    client: TestClient, superuser_token_headers: dict[str, str]
) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "字段编码对照表"
    ws.append(["字段编码", "字段名称", "所属分组"])
    ws.append(["sms_signature", "短信签名", "签名与模板"])
    ws.append(["main_port_number", "主端口号", "端口信息"])
    output = io.BytesIO()
    wb.save(output)

    r = client.post(
        f"{settings.API_V1_STR}/export-groups/import",
        headers=superuser_token_headers,
        files={
            "file": (
                "字段编码对照表.xlsx",
                output.getvalue(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert r.status_code == 200
    data = r.json()
    assert data["success_count"] == 1
    assert data["group_name"] == "字段编码对照表"
    assert data["field_count"] == 2


def test_import_export_group_invalid_field_code(
    client: TestClient, superuser_token_headers: dict[str, str]
) -> None:
    content = _build_import_xlsx(
        [
            ["导入字段组B", "sms_signature", "短信签名", 0, "是"],
            ["导入字段组B", "not_a_real_field", "不存在", 1, "是"],
        ]
    )
    r = client.post(
        f"{settings.API_V1_STR}/export-groups/import",
        headers=superuser_token_headers,
        files={
            "file": (
                "group.xlsx",
                content,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert r.status_code == 200
    data = r.json()
    assert data["success_count"] == 0
    assert data["error_count"] == 1
    assert data["errors"][0]["value"] == "not_a_real_field"


def test_import_export_group_bad_extension(
    client: TestClient, superuser_token_headers: dict[str, str]
) -> None:
    r = client.post(
        f"{settings.API_V1_STR}/export-groups/import",
        headers=superuser_token_headers,
        files={"file": ("group.txt", b"not excel", "text/plain")},
    )
    assert r.status_code == 400
    assert "仅支持" in r.json()["detail"]


def test_delete_export_group_referenced_by_filing_task(
    client: TestClient,
    superuser_token_headers: dict[str, str],
    db: Session,
) -> None:
    gid = _create_group(client, superuser_token_headers, name="历史字段组")
    user = db.exec(select(User)).first()
    assert user is not None

    task = FilingTask(
        task_name=f"BEI-TEST-{uuid.uuid4()}",
        qualification_ids=[],
        port_ids=[],
        export_group_id=uuid.UUID(gid),
        export_group_name="历史字段组",
        qualification_count=0,
        port_count=0,
        operator_id=user.id,
    )
    db.add(task)
    db.commit()

    # 删除字段组不再受报备任务限制
    r = client.delete(
        f"{settings.API_V1_STR}/export-groups/{gid}",
        headers=superuser_token_headers,
    )
    assert r.status_code == 200

    # 任务仍在，引用被置空，展示用快照组名
    db.expire_all()
    task_after = db.get(FilingTask, task.id)
    assert task_after is not None
    assert task_after.export_group_id is None

    detail = client.get(
        f"{settings.API_V1_STR}/filing-tasks/{task.id}",
        headers=superuser_token_headers,
    ).json()
    assert detail["export_group_name"] == "历史字段组"


def test_regenerate_filing_task_after_group_deleted(
    client: TestClient,
    superuser_token_headers: dict[str, str],
    db: Session,
) -> None:
    gid = _create_group(client, superuser_token_headers)
    user = db.exec(select(User)).first()
    assert user is not None

    task = FilingTask(
        task_name=f"BEI-TEST-{uuid.uuid4()}",
        qualification_ids=[],
        port_ids=[],
        export_group_id=uuid.UUID(gid),
        qualification_count=0,
        port_count=0,
        operator_id=user.id,
    )
    db.add(task)
    db.commit()

    r = client.delete(
        f"{settings.API_V1_STR}/export-groups/{gid}",
        headers=superuser_token_headers,
    )
    assert r.status_code == 200

    r = client.post(
        f"{settings.API_V1_STR}/filing-tasks/{task.id}/regenerate",
        headers=superuser_token_headers,
    )
    assert r.status_code == 400
    assert "已被删除" in r.json()["detail"]


def test_delete_export_group_success(
    client: TestClient, superuser_token_headers: dict[str, str]
) -> None:
    gid = _create_group(client, superuser_token_headers)
    r = client.delete(
        f"{settings.API_V1_STR}/export-groups/{gid}",
        headers=superuser_token_headers,
    )
    assert r.status_code == 200
