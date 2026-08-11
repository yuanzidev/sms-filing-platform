"""Regression tests for filing task export — all selected fields must appear in Excel."""

import hashlib
from io import BytesIO
from typing import Generator
import uuid
import zipfile

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from PIL import Image as PILImage
from sqlmodel import Session, delete

from app.core.config import settings
from app.core.db import engine
from app.models import FileAttachment, FilingSubPortUsage, FilingTask
from app.services.export_field_registry import REGISTRY


@pytest.fixture(scope="module", autouse=True)
def _cleanup_filing_tasks() -> Generator[None, None, None]:
    yield
    with Session(engine) as session:
        # 占用记录引用 user（operator_id），须先清理，否则 session 级 user 清理会违反外键
        session.execute(delete(FilingSubPortUsage))
        session.execute(delete(FilingTask))
        session.commit()


def _create_qualification(client, headers, name):
    r = client.post(
        f"{settings.API_V1_STR}/qualifications",
        headers=headers,
        json={
            "enterprise_name": name,
            "sms_signature": "签名X",
            "signature_type": "自营签名",
            "specific_usage": "用户登录",
            "diversion_number": "13800000000",
            "link_address": "https://example.com",
            "legal_representative_cert_type": "身份证",
            "legal_representative_cert_number": "110101199001011234",
            "legal_representative_cert_address": "北京市朝阳区",
        },
    )
    assert r.status_code == 200, r.text
    return r.json()["id"]


def _create_port(client, headers, main_port_number):
    r = client.post(
        f"{settings.API_V1_STR}/port-info",
        headers=headers,
        json={
            "carrier": "中国移动",
            "main_port_number": main_port_number,
            "enterprise_name": "测试企业",
            "group_code": "G001",
            "carrier_room": "机房A",
            "enterprise_room": "机房B",
            "port_type": "短信",
            "operation_type": "新增",
            "authorization_letter": "授字001",
        },
    )
    assert r.status_code == 200, r.text
    return r.json()["id"]


def _create_export_group_all_fields(client, headers, name):
    fields = [
        {"field_name": f.name, "field_label": f.label, "sort_order": i}
        for i, f in enumerate(REGISTRY, 1)
    ]
    r = client.post(
        f"{settings.API_V1_STR}/export-groups",
        headers=headers,
        json={"name": name, "fields": fields},
    )
    assert r.status_code == 200, r.text
    return r.json()["id"]


def _create_export_group(client, headers, name):
    r = client.post(
        f"{settings.API_V1_STR}/export-groups",
        headers=headers,
        json={
            "name": name,
            "fields": [
                {
                    "field_name": "main_port_number",
                    "field_label": "主端口号",
                    "sort_order": 1,
                },
            ],
        },
    )
    assert r.status_code == 200, r.text
    return r.json()["id"]


def _png_bytes(color: str) -> bytes:
    image = PILImage.new("RGB", (20, 20), color=color)
    buf = BytesIO()
    image.save(buf, format="PNG")
    return buf.getvalue()


def test_export_includes_all_registry_fields(
    client: TestClient, superuser_token_headers: dict[str, str]
) -> None:
    qual_id = _create_qualification(client, superuser_token_headers, "全字段企业")
    port_id = _create_port(client, superuser_token_headers, "10698全部")
    group_id = _create_export_group_all_fields(
        client, superuser_token_headers, "全字段组"
    )

    r = client.post(
        f"{settings.API_V1_STR}/filing-tasks",
        headers=superuser_token_headers,
        json={
            "qualification_ids": [qual_id],
            "port_ids": [port_id],
            "export_group_id": group_id,
        },
    )
    assert r.status_code == 200, r.text
    task_id = r.json()["id"]

    r = client.get(
        f"{settings.API_V1_STR}/filing-tasks/{task_id}/download",
        headers=superuser_token_headers,
    )
    assert r.status_code == 200

    wb = load_workbook(BytesIO(r.content))
    ws = wb.active
    header_row = [c.value for c in ws[1]]

    expected_labels = {f.label for f in REGISTRY}
    actual_labels = set(header_row)
    missing = expected_labels - actual_labels
    assert not missing, f"导出缺失列: {missing}"


def test_filing_task_download_embeds_qualification_and_port_images(
    client: TestClient, superuser_token_headers: dict[str, str], monkeypatch
) -> None:
    from app.api.routes import filing_tasks

    qual_id = _create_qualification(client, superuser_token_headers, "图片导出企业")
    port_id = _create_port(client, superuser_token_headers, "10698IMG")
    qual_png = _png_bytes("red")
    port_png = _png_bytes("blue")

    class _MemoryStorage:
        def __init__(self) -> None:
            self.files = {
                "qualification-cert.png": qual_png,
                "port-auth.png": port_png,
            }

        def upload(self, key, data, content_type):
            self.files[key] = data
            return key

        def download(self, key):
            return self.files[key]

        def exists(self, key):
            return key in self.files

        def delete(self, key):
            self.files.pop(key, None)

    storage = _MemoryStorage()
    monkeypatch.setattr(filing_tasks, "get_storage", lambda: storage)

    with Session(engine) as session:
        session.add_all(
            [
                FileAttachment(
                    original_name="image_row1_col4.png",
                    stored_path="qualification-cert.png",
                    file_size=len(qual_png),
                    mime_type="image/png",
                    md5_hash=hashlib.md5(qual_png).hexdigest(),
                    entity_type="qualification_info",
                    entity_id=uuid.UUID(qual_id),
                    field_name=None,
                ),
                FileAttachment(
                    original_name="auth.png",
                    stored_path="port-auth.png",
                    file_size=len(port_png),
                    mime_type="image/png",
                    md5_hash=hashlib.md5(port_png).hexdigest(),
                    entity_type="port_info",
                    entity_id=uuid.UUID(port_id),
                    field_name="授权书图片",
                ),
            ]
        )
        session.commit()

    r = client.post(
        f"{settings.API_V1_STR}/export-groups",
        headers=superuser_token_headers,
        json={
            "name": "图片导出组",
            "fields": [
                {
                    "field_name": "cert_image",
                    "field_label": "单位证件图片",
                    "sort_order": 1,
                },
                {
                    "field_name": "auth_image",
                    "field_label": "授权书图片",
                    "sort_order": 2,
                },
            ],
        },
    )
    assert r.status_code == 200, r.text
    group_id = r.json()["id"]

    r = client.post(
        f"{settings.API_V1_STR}/filing-tasks",
        headers=superuser_token_headers,
        json={
            "qualification_ids": [qual_id],
            "port_ids": [port_id],
            "export_group_id": group_id,
        },
    )
    assert r.status_code == 200, r.text
    task_id = r.json()["id"]

    r = client.get(
        f"{settings.API_V1_STR}/filing-tasks/{task_id}/download",
        headers=superuser_token_headers,
    )
    assert r.status_code == 200

    with zipfile.ZipFile(BytesIO(r.content)) as zf:
        media_files = [name for name in zf.namelist() if name.startswith("xl/media/")]
        drawing_files = [
            name for name in zf.namelist() if name.startswith("xl/drawings/")
        ]
    assert len(media_files) >= 2
    assert drawing_files


def test_filing_task_image_columns_show_no_image_when_attachment_missing(
    client: TestClient, superuser_token_headers: dict[str, str]
) -> None:
    qual_id = _create_qualification(client, superuser_token_headers, "无图片企业")
    port_id = _create_port(client, superuser_token_headers, "10698NOIMG")

    r = client.post(
        f"{settings.API_V1_STR}/export-groups",
        headers=superuser_token_headers,
        json={
            "name": "无图片字段组",
            "fields": [
                {"field_name": "cert_image", "field_label": "单位证件图片", "sort_order": 1},
                {"field_name": "auth_image", "field_label": "授权书图片", "sort_order": 2},
            ],
        },
    )
    assert r.status_code == 200, r.text
    group_id = r.json()["id"]

    r = client.post(
        f"{settings.API_V1_STR}/filing-tasks",
        headers=superuser_token_headers,
        json={
            "qualification_ids": [qual_id],
            "port_ids": [port_id],
            "export_group_id": group_id,
        },
    )
    assert r.status_code == 200, r.text
    task_id = r.json()["id"]

    r = client.get(
        f"{settings.API_V1_STR}/filing-tasks/{task_id}/download",
        headers=superuser_token_headers,
    )
    assert r.status_code == 200

    wb = load_workbook(BytesIO(r.content))
    ws = wb.active
    assert ws.cell(row=2, column=1).value == "无图片"
    assert ws.cell(row=2, column=2).value == "无图片"


def test_create_filing_task_with_auto_sub_ports(
    client: TestClient, superuser_token_headers: dict[str, str]
) -> None:
    """自动分配模式下，导出含随机生成的子端口号"""
    qual_id = _create_qualification(client, superuser_token_headers, "自动分配企业")
    # 主端口行（sub_port_number 为空）
    port_id = _create_port(client, superuser_token_headers, "10698AUTO")

    # 仅勾选主端口号、子端口号两列的字段组
    r = client.post(
        f"{settings.API_V1_STR}/export-groups",
        headers=superuser_token_headers,
        json={
            "name": "子端口导出组",
            "fields": [
                {
                    "field_name": "main_port_number",
                    "field_label": "主端口号",
                    "sort_order": 1,
                },
                {
                    "field_name": "sub_port_number",
                    "field_label": "子端口号",
                    "sort_order": 2,
                },
            ],
        },
    )
    group_id = r.json()["id"]

    r = client.post(
        f"{settings.API_V1_STR}/filing-tasks",
        headers=superuser_token_headers,
        json={
            "qualification_ids": [qual_id],
            "port_ids": [port_id],
            "export_group_id": group_id,
            "auto_allocate_sub_ports": True,
            "sub_port_range_start": 300001,
            "sub_port_range_end": 300100,
        },
    )
    assert r.status_code == 200, r.text
    task_id = r.json()["id"]

    r = client.get(
        f"{settings.API_V1_STR}/filing-tasks/{task_id}/download",
        headers=superuser_token_headers,
    )
    wb = load_workbook(BytesIO(r.content))
    ws = wb.active
    # 第 2 行第 2 列是子端口号
    sub_port_value = ws.cell(row=2, column=2).value
    assert sub_port_value is not None
    assert str(sub_port_value).startswith("3000")  # 在范围内


def test_create_filing_task_range_exhausted_409(
    client: TestClient, superuser_token_headers: dict[str, str]
) -> None:
    """范围耗尽时 409"""
    qual_id = _create_qualification(client, superuser_token_headers, "范围耗尽企业")
    # 创建 2 个资质，但范围只够 1 个
    qual_id_2 = _create_qualification(client, superuser_token_headers, "范围耗尽企业2")
    port_id = _create_port(client, superuser_token_headers, "10698EXH")

    r = client.post(
        f"{settings.API_V1_STR}/export-groups",
        headers=superuser_token_headers,
        json={
            "name": "耗尽组",
            "fields": [
                {
                    "field_name": "main_port_number",
                    "field_label": "主端口号",
                    "sort_order": 1,
                },
            ],
        },
    )
    group_id = r.json()["id"]

    r = client.post(
        f"{settings.API_V1_STR}/filing-tasks",
        headers=superuser_token_headers,
        json={
            "qualification_ids": [qual_id, qual_id_2],
            "port_ids": [port_id],
            "export_group_id": group_id,
            "auto_allocate_sub_ports": True,
            "sub_port_range_start": 400001,
            "sub_port_range_end": 400001,  # 只 1 个号码，需要 2 个
        },
    )
    assert r.status_code == 409
    assert "10698EXH" in r.json()["detail"]


def test_auto_sub_ports_range_too_large_rejected(
    client: TestClient, superuser_token_headers: dict[str, str]
) -> None:
    """范围超过 100 万 → 400"""
    qual_id = _create_qualification(client, superuser_token_headers, "范围过大企业")
    port_id = _create_port(client, superuser_token_headers, "10698TOOBIG")
    group_id = _create_export_group_all_fields(
        client, superuser_token_headers, "范围过大组"
    )

    r = client.post(
        f"{settings.API_V1_STR}/filing-tasks",
        headers=superuser_token_headers,
        json={
            "qualification_ids": [qual_id],
            "port_ids": [port_id],
            "export_group_id": group_id,
            "auto_allocate_sub_ports": True,
            "sub_port_range_start": 100000,
            "sub_port_range_end": 2000000,
        },
    )
    assert r.status_code == 400
    assert "范围过大" in r.json()["detail"]


def test_auto_sub_ports_non_6digit_rejected(
    client: TestClient, superuser_token_headers: dict[str, str]
) -> None:
    """范围不是 6 位数字 → 400"""
    qual_id = _create_qualification(client, superuser_token_headers, "非6位企业")
    port_id = _create_port(client, superuser_token_headers, "10698NOT6")
    group_id = _create_export_group_all_fields(
        client, superuser_token_headers, "非6位组"
    )

    r = client.post(
        f"{settings.API_V1_STR}/filing-tasks",
        headers=superuser_token_headers,
        json={
            "qualification_ids": [qual_id],
            "port_ids": [port_id],
            "export_group_id": group_id,
            "auto_allocate_sub_ports": True,
            "sub_port_range_start": 10,
            "sub_port_range_end": 99,
        },
    )
    assert r.status_code == 400
    assert "6位" in r.json()["detail"]


def test_failed_upload_releases_sub_ports(
    client: TestClient, superuser_token_headers: dict[str, str], monkeypatch
) -> None:
    """文件上传失败时，已分配的子端口占用被释放，不残留"""
    from sqlmodel import select

    from app.core.db import engine
    from app.models import FilingSubPortUsage

    qual_id = _create_qualification(client, superuser_token_headers, "上传失败企业")
    port_id = _create_port(client, superuser_token_headers, "10698UPLFAIL")

    r = client.post(
        f"{settings.API_V1_STR}/export-groups",
        headers=superuser_token_headers,
        json={
            "name": "上传失败组",
            "fields": [
                {
                    "field_name": "main_port_number",
                    "field_label": "主端口号",
                    "sort_order": 1,
                },
                {
                    "field_name": "sub_port_number",
                    "field_label": "子端口号",
                    "sort_order": 2,
                },
            ],
        },
    )
    group_id = r.json()["id"]

    from app.api.routes import filing_tasks

    class _FailingStorage:
        def download(self, path):
            raise RuntimeError("storage down")

        def upload(self, key, data, content_type):
            raise RuntimeError("storage down")

        def delete(self, path):
            pass

    monkeypatch.setattr(filing_tasks, "get_storage", lambda: _FailingStorage())

    r = client.post(
        f"{settings.API_V1_STR}/filing-tasks",
        headers=superuser_token_headers,
        json={
            "qualification_ids": [qual_id],
            "port_ids": [port_id],
            "export_group_id": group_id,
            "auto_allocate_sub_ports": True,
            "sub_port_range_start": 600001,
            "sub_port_range_end": 600100,
        },
    )
    assert r.status_code == 500

    # 失败路径不残留占用记录
    with Session(engine) as session:
        stmt = select(FilingSubPortUsage).where(
            FilingSubPortUsage.main_port_number == "10698UPLFAIL"
        )
        usages = list(session.exec(stmt).all())
        assert len(usages) == 0


def test_delete_filing_task_keeps_usage(
    client: TestClient, superuser_token_headers: dict[str, str]
) -> None:
    """删除报备任务后，占用记录仍存在"""
    from sqlmodel import Session, select

    from app.core.db import engine
    from app.models import FilingSubPortUsage

    qual_id = _create_qualification(client, superuser_token_headers, "保留占用企业")
    port_id = _create_port(client, superuser_token_headers, "10698KEEP")

    r = client.post(
        f"{settings.API_V1_STR}/export-groups",
        headers=superuser_token_headers,
        json={
            "name": "保留组",
            "fields": [
                {
                    "field_name": "main_port_number",
                    "field_label": "主端口号",
                    "sort_order": 1,
                },
                {
                    "field_name": "sub_port_number",
                    "field_label": "子端口号",
                    "sort_order": 2,
                },
            ],
        },
    )
    group_id = r.json()["id"]

    r = client.post(
        f"{settings.API_V1_STR}/filing-tasks",
        headers=superuser_token_headers,
        json={
            "qualification_ids": [qual_id],
            "port_ids": [port_id],
            "export_group_id": group_id,
            "auto_allocate_sub_ports": True,
            "sub_port_range_start": 500001,
            "sub_port_range_end": 500100,
        },
    )
    task_id = r.json()["id"]

    # 删除报备任务
    r = client.delete(
        f"{settings.API_V1_STR}/filing-tasks/{task_id}",
        headers=superuser_token_headers,
    )
    assert r.status_code == 200

    # 占用记录仍在，filing_task_id 变 None
    with Session(engine) as session:
        stmt = select(FilingSubPortUsage).where(
            FilingSubPortUsage.main_port_number == "10698KEEP"
        )
        usages = list(session.exec(stmt).all())
        assert len(usages) >= 1
        for u in usages:
            assert u.filing_task_id is None


def test_sub_port_availability(
    client: TestClient, superuser_token_headers: dict[str, str]
) -> None:
    """子端口可用性查询：已占用的号码计入 used，available = total - used"""
    from sqlmodel import select

    from app.core.db import engine
    from app.models import FilingSubPortUsage, User

    # 先插入 2 条占用记录，确保 used > 0
    with Session(engine) as session:
        operator = session.exec(
            select(User).where(User.email == settings.FIRST_SUPERUSER)
        ).first()
        assert operator is not None
        session.add_all(
            [
                FilingSubPortUsage(
                    main_port_number="10698AVAIL",
                    port_number="100003",
                    operator_id=operator.id,
                ),
                FilingSubPortUsage(
                    main_port_number="10698AVAIL",
                    port_number="100005",
                    operator_id=operator.id,
                ),
            ]
        )
        session.commit()

    r = client.get(
        f"{settings.API_V1_STR}/filing-tasks/sub-port-availability",
        headers=superuser_token_headers,
        params={
            "main_port_numbers": "10698AVAIL",
            "range_start": 100001,
            "range_end": 100010,
        },
    )
    assert r.status_code == 200
    data = r.json()
    assert "10698AVAIL" in data
    info = data["10698AVAIL"]
    assert info["total"] == 10
    assert info["used"] == 2
    assert info["available"] == 8
    assert info["available"] + info["used"] == info["total"]


def test_search_filing_tasks_by_operator_name(
    client: TestClient, superuser_token_headers: dict[str, str]
) -> None:
    """关键词应匹配操作人姓名"""
    qual_id = _create_qualification(client, superuser_token_headers, "搜索企业")
    port_id = _create_port(client, superuser_token_headers, "10698SCH")
    group_id = _create_export_group(client, superuser_token_headers, "搜索字段组")

    r = client.post(
        f"{settings.API_V1_STR}/filing-tasks",
        headers=superuser_token_headers,
        json={
            "qualification_ids": [qual_id],
            "port_ids": [port_id],
            "export_group_id": group_id,
        },
    )
    assert r.status_code == 200

    # 用操作人（admin用户）的部分名称搜索
    r = client.get(
        f"{settings.API_V1_STR}/filing-tasks",
        headers=superuser_token_headers,
        params={"keyword": "admin"},
    )
    assert r.status_code == 200
    assert r.json()["total"] >= 1


def test_create_filing_task_with_custom_name(
    client: TestClient, superuser_token_headers: dict[str, str]
) -> None:
    qual_id = _create_qualification(client, superuser_token_headers, "自定义名企业")
    port_id = _create_port(client, superuser_token_headers, "10698CSTM")
    group_id = _create_export_group(client, superuser_token_headers, "自定义名组")

    r = client.post(
        f"{settings.API_V1_STR}/filing-tasks",
        headers=superuser_token_headers,
        json={
            "qualification_ids": [qual_id],
            "port_ids": [port_id],
            "export_group_id": group_id,
            "task_name": "我的自定义任务名",
        },
    )
    assert r.status_code == 200
    assert r.json()["task_name"] == "我的自定义任务名"
