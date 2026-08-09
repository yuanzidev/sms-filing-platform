"""Tests for port-info API: page_size upper bound, keyword/city/port_type/main_port_number filters."""
import uuid

from fastapi.testclient import TestClient

from app.core.config import settings


def test_port_info_accepts_large_page_size(
    client: TestClient, superuser_token_headers: dict[str, str]
) -> None:
    r = client.get(
        f"{settings.API_V1_STR}/port-info",
        headers=superuser_token_headers,
        params={"page_size": 500},
    )
    assert r.status_code == 200, r.text
    body = r.json()
    assert body["page_size"] == 500


def test_port_info_rejects_too_large_page_size(
    client: TestClient, superuser_token_headers: dict[str, str]
) -> None:
    r = client.get(
        f"{settings.API_V1_STR}/port-info",
        headers=superuser_token_headers,
        params={"page_size": 501},
    )
    assert r.status_code == 422


def test_import_port_info_with_empty_operation_and_group(
    client: TestClient, superuser_token_headers: dict[str, str]
) -> None:
    """操作类型/集团编码为空可导入"""
    from io import BytesIO
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    headers = [
        "运营商", "主端口号", "主端口备案公司", "端口类型",
        "运营商接入机房及设备", "企业接入机房及设备", "授权书",
    ]
    for col_idx, h in enumerate(headers, 1):
        ws.cell(row=1, column=col_idx, value=h)
    ws.cell(row=2, column=1, value="中国移动")
    ws.cell(row=2, column=2, value="10698999")
    ws.cell(row=2, column=3, value="测试企业")
    ws.cell(row=2, column=4, value="短信")
    ws.cell(row=2, column=5, value="机房A")
    ws.cell(row=2, column=6, value="机房B")
    ws.cell(row=2, column=7, value="授字001")

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    r = client.post(
        f"{settings.API_V1_STR}/port-info/import",
        headers=superuser_token_headers,
        files={"file": ("test.xlsx", buf, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert r.status_code == 200, r.text
    body = r.json()
    assert body["success_count"] >= 1


def test_import_port_infos_collects_errors_and_writes_valid_rows(
    client: TestClient, superuser_token_headers: dict[str, str]
) -> None:
    """混合文件：错误行收集所有错误（必填缺失+日期无效），有效行批量写入"""
    from io import BytesIO
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    headers = ["运营商", "主端口号", "主端口备案公司", "端口类型", "端口入网时间"]
    for col_idx, h in enumerate(headers, 1):
        ws.cell(row=1, column=col_idx, value=h)
    # 行2：运营商为空
    ws.cell(row=2, column=2, value="10690001")
    ws.cell(row=2, column=3, value="错误企业A")
    ws.cell(row=2, column=4, value="短信")
    # 行3：日期格式无效
    ws.cell(row=3, column=1, value="中国移动")
    ws.cell(row=3, column=2, value="10690002")
    ws.cell(row=3, column=3, value="错误企业B")
    ws.cell(row=3, column=4, value="短信")
    ws.cell(row=3, column=5, value="2024-13-99")
    # 行4：有效
    ws.cell(row=4, column=1, value="中国移动")
    ws.cell(row=4, column=2, value="10690003")
    ws.cell(row=4, column=3, value="有效企业C")
    ws.cell(row=4, column=4, value="短信")

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    r = client.post(
        f"{settings.API_V1_STR}/port-info/import",
        headers=superuser_token_headers,
        files={"file": ("test.xlsx", buf, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert r.status_code == 200, r.text
    body = r.json()
    assert body["success_count"] == 1
    assert body["error_count"] == 2
    err_fields = {e["field"] for e in body["errors"]}
    assert err_fields == {"运营商", "端口入网时间"}


def test_port_info_import_error_report_xlsx(
    client: TestClient, superuser_token_headers: dict[str, str]
) -> None:
    from io import BytesIO
    from openpyxl import load_workbook

    errors = [
        {"row": 2, "field": "运营商", "value": "", "reason": "运营商不能为空", "suggestion": "请填写运营商"},
    ]
    r = client.post(
        f"{settings.API_V1_STR}/port-info/import/error-report",
        headers=superuser_token_headers,
        json={"errors": errors},
    )
    assert r.status_code == 200
    assert "spreadsheetml" in r.headers["content-type"]
    wb = load_workbook(BytesIO(r.content))
    ws = wb.active
    assert ws.title == "导入错误报告"
    assert ws.cell(row=2, column=2).value == "运营商"


def test_filter_port_infos_by_keyword(
    client: TestClient, superuser_token_headers: dict[str, str]
) -> None:
    marker = uuid.uuid4().hex
    sub_port = f"9999{marker[:4]}"
    # 创建一条数据
    client.post(
        f"{settings.API_V1_STR}/port-info",
        headers=superuser_token_headers,
        json={
            "carrier": "中国移动",
            "main_port_number": f"1069{marker[:6]}",
            "enterprise_name": f"测试备案公司{marker}",
            "port_type": "短信",
            "carrier_room": "X机房",
            "enterprise_room": "Y机房",
            "authorization_letter": "授字001",
            "sub_port_number": sub_port,
        },
    )
    # keyword 命中子端口号
    r = client.get(
        f"{settings.API_V1_STR}/port-info",
        headers=superuser_token_headers,
        params={"keyword": sub_port},
    )
    assert r.status_code == 200
    body = r.json()
    assert body["total"] >= 1

    # keyword 命中企业名称
    r2 = client.get(
        f"{settings.API_V1_STR}/port-info",
        headers=superuser_token_headers,
        params={"keyword": marker},
    )
    assert r2.json()["total"] >= 1

    # keyword 无命中时返回 0(验证 keyword 过滤真实生效)
    r3 = client.get(
        f"{settings.API_V1_STR}/port-info",
        headers=superuser_token_headers,
        params={"keyword": f"nomatch{marker}"},
    )
    assert r3.status_code == 200
    assert r3.json()["total"] == 0

    # main_port_number 参数过滤
    r4 = client.get(
        f"{settings.API_V1_STR}/port-info",
        headers=superuser_token_headers,
        params={"main_port_number": f"1069{marker[:6]}"},
    )
    assert r4.status_code == 200
    assert r4.json()["total"] >= 1


def test_filter_port_infos_by_city_and_type(
    client: TestClient, superuser_token_headers: dict[str, str]
) -> None:
    marker = uuid.uuid4().hex
    city = f"城市{marker[:4]}"
    client.post(
        f"{settings.API_V1_STR}/port-info",
        headers=superuser_token_headers,
        json={
            "carrier": "中国联通",
            "main_port_number": f"1069{marker[:6]}",
            "enterprise_name": f"城市测试公司{marker}",
            "port_type": "语音",
            "city": city,
            "carrier_room": "A机房",
            "enterprise_room": "B机房",
            "authorization_letter": "授字002",
        },
    )
    r = client.get(
        f"{settings.API_V1_STR}/port-info",
        headers=superuser_token_headers,
        params={"city": city, "port_type": "语音"},
    )
    assert r.status_code == 200
    body = r.json()
    assert body["total"] >= 1

    # city 无命中时返回 0(验证 city 过滤真实生效)
    r2 = client.get(
        f"{settings.API_V1_STR}/port-info",
        headers=superuser_token_headers,
        params={"city": f"不存在{marker}"},
    )
    assert r2.status_code == 200
    assert r2.json()["total"] == 0
