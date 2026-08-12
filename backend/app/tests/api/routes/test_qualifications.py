"""Tests for qualifications API: updated field schema."""
from fastapi.testclient import TestClient

from app.core.config import settings


def test_list_qualifications_filter_by_sms_signature(
    client: TestClient, superuser_token_headers: dict[str, str]
) -> None:
    for sig in ("平台A签名", "平台B签名"):
        client.post(
            f"{settings.API_V1_STR}/qualifications",
            headers=superuser_token_headers,
            json={
                "enterprise_name": f"测试企业 {sig}",
                "sms_signature": sig,
                "legal_representative_cert_type": "身份证",
                "legal_representative_cert_number": "110101199001011234",
                "legal_representative_cert_address": "北京市朝阳区XX路1号",
            },
        )
    r = client.get(
        f"{settings.API_V1_STR}/qualifications",
        headers=superuser_token_headers,
        params={"sms_signature": "平台A"},
    )
    assert r.status_code == 200
    body = r.json()
    assert body["total"] >= 1
    assert all("平台A" in item["sms_signature"] for item in body["data"])


def test_list_qualifications_filter_by_identity_cert_number(
    client: TestClient, superuser_token_headers: dict[str, str]
) -> None:
    marker = "IDSEARCH001"
    legal_number = f"11010119900101{marker}"
    responsible_number = f"11010119900202{marker}"
    handler_number = f"11010119900303{marker}"

    for enterprise_name, field_name, value in (
        ("法人证件搜索企业", "legal_representative_cert_number", legal_number),
        ("责任人证件搜索企业", "responsible_cert_number", responsible_number),
        ("经办人证件搜索企业", "handler_cert_number", handler_number),
    ):
        payload = {
            "enterprise_name": enterprise_name,
            "cert_number": f"91330100{marker}{field_name}",
            field_name: value,
        }
        r = client.post(
            f"{settings.API_V1_STR}/qualifications",
            headers=superuser_token_headers,
            json=payload,
        )
        assert r.status_code == 200, r.text

    r = client.get(
        f"{settings.API_V1_STR}/qualifications",
        headers=superuser_token_headers,
        params={"identity_cert_number": marker},
    )
    assert r.status_code == 200
    names = {item["enterprise_name"] for item in r.json()["data"]}
    assert {
        "法人证件搜索企业",
        "责任人证件搜索企业",
        "经办人证件搜索企业",
    }.issubset(names)


def test_template_has_required_headers(
    client: TestClient, superuser_token_headers: dict[str, str]
) -> None:
    from io import BytesIO
    from openpyxl import load_workbook

    r = client.get(
        f"{settings.API_V1_STR}/qualifications/template", headers=superuser_token_headers
    )
    assert r.status_code == 200
    wb = load_workbook(BytesIO(r.content))
    ws = wb.active
    headers = [c.value for c in ws[1]]
    assert "企业名称" in headers
    assert "法人证件类型" in headers
    assert "法人证件号码" in headers
    assert "法人证件地址" in headers
    assert "单位证件图片" in headers
    assert "签名" not in headers


def _build_xlsx(headers: list[str], rows: list[list]) -> bytes:
    from io import BytesIO
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.append(headers)
    for r in rows:
        ws.append(r)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_import_rejects_missing_enterprise_name(
    client: TestClient, superuser_token_headers: dict[str, str]
) -> None:
    """缺少「企业名称」列视为模板不匹配，直接 400"""
    headers = ["法人证件类型", "法人证件号码"]
    rows = [["身份证", "110101199001011234"]]
    data = _build_xlsx(headers, rows)

    files = {"file": ("test.xlsx", data, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
    r = client.post(
        f"{settings.API_V1_STR}/qualifications/import",
        headers=superuser_token_headers,
        files=files,
    )
    assert r.status_code == 400
    assert "企业名称" in r.json()["detail"]


def test_import_succeeds_without_legal_rep_columns(
    client: TestClient, superuser_token_headers: dict[str, str]
) -> None:
    """法人证件列完全缺失时仍可导入（法人字段选填）"""
    headers = ["企业名称", "短信签名"]
    rows = [["测试企业C", "测试签名"]]
    data = _build_xlsx(headers, rows)

    files = {"file": ("test.xlsx", data, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
    r = client.post(
        f"{settings.API_V1_STR}/qualifications/import",
        headers=superuser_token_headers,
        files=files,
    )
    assert r.status_code == 200
    assert r.json()["success_count"] == 1


def test_import_success_with_required_fields(
    client: TestClient, superuser_token_headers: dict[str, str]
) -> None:
    headers = [
        "企业名称", "法人证件类型", "法人证件号码", "法人证件地址",
    ]
    rows = [["测试企业B", "身份证", "110101199001011234", "北京市朝阳区XX路1号"]]
    data = _build_xlsx(headers, rows)

    files = {"file": ("test.xlsx", data, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
    r = client.post(
        f"{settings.API_V1_STR}/qualifications/import",
        headers=superuser_token_headers,
        files=files,
    )
    assert r.status_code == 200
    assert r.json()["success_count"] == 1


def test_template_column_order_matches_new_spec(
    client: TestClient, superuser_token_headers: dict[str, str]
) -> None:
    from io import BytesIO
    from openpyxl import load_workbook

    r = client.get(
        f"{settings.API_V1_STR}/qualifications/template", headers=superuser_token_headers
    )
    assert r.status_code == 200
    wb = load_workbook(BytesIO(r.content))
    ws = wb.active
    headers = [c.value for c in ws[1]]
    # 新模板 45 列关键位置断言（经办人字段前置、法人字段跟随）
    assert headers[11] == "经办人姓名", f"col12 应为「经办人姓名」，实际：{headers[11]}"
    assert headers[17] == "法人姓名", f"col18 应为「法人姓名」，实际：{headers[17]}"
    assert headers[21] == "法人身份证正面", f"col22 应为「法人身份证正面」，实际：{headers[21]}"
    assert headers[22] == "法人身份证反面", f"col23 应为「法人身份证反面」，实际：{headers[22]}"
    assert headers[36] == "引流链接", f"col37 应为「引流链接」，实际：{headers[36]}"
    assert headers[41] == "签名举证附件", f"col42 应为「签名举证附件」，实际：{headers[41]}"
    assert headers[42] == "引流号码举证附件", f"col43 应为「引流号码举证附件」，实际：{headers[42]}"
    assert headers[43] == "引流链接举证", f"col44 应为「引流链接举证」，实际：{headers[43]}"
    # 旧名不应存在
    assert "链接地址" not in headers
    assert "经办人身份证正面" not in headers
    assert "经办人身份证反面" not in headers
    assert "引流举证附件" not in headers
    # 总列数
    assert len([h for h in headers if h]) == 45


def test_import_accepts_renamed_link_address_header(
    client: TestClient, superuser_token_headers: dict[str, str]
) -> None:
    headers = [
        "企业名称", "法人证件类型", "法人证件号码", "法人证件地址", "引流链接", "短信签名",
    ]
    rows = [["测试企业链接", "身份证", "110101199001011234", "北京市朝阳区XX路1号", "https://example.com", "【测试签名】"]]
    data = _build_xlsx(headers, rows)

    files = {"file": ("test.xlsx", data, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
    r = client.post(
        f"{settings.API_V1_STR}/qualifications/import",
        headers=superuser_token_headers,
        files=files,
    )
    assert r.status_code == 200
    assert r.json()["success_count"] == 1
    # 验证值确实落到 link_address 和 sms_signature 字段
    list_r = client.get(
        f"{settings.API_V1_STR}/qualifications",
        headers=superuser_token_headers,
        params={"enterprise_name": "测试企业链接"},
    )
    assert list_r.status_code == 200
    item = list_r.json()["data"][0]
    assert item["link_address"] == "https://example.com"
    assert item["sms_signature"] == "【测试签名】"


def test_import_qualifications_with_empty_legal_fields(
    client: TestClient, superuser_token_headers: dict[str, str]
) -> None:
    """法人证件类型/号码/地址为空可导入"""
    from io import BytesIO
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    headers = ["企业名称", "法人证件类型", "法人证件号码", "法人证件地址", "短信签名"]
    for col_idx, h in enumerate(headers, 1):
        ws.cell(row=1, column=col_idx, value=h)
    # 法人字段留空
    ws.cell(row=2, column=1, value="测试企业")
    ws.cell(row=2, column=5, value="测试签名")

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    r = client.post(
        f"{settings.API_V1_STR}/qualifications/import",
        headers=superuser_token_headers,
        files={"file": ("test.xlsx", buf, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert r.status_code == 200, r.text
    body = r.json()
    assert body["success_count"] >= 1


def test_qualification_template_notes_mention_optional_legal(
    client: TestClient, superuser_token_headers: dict[str, str]
) -> None:
    from io import BytesIO
    from openpyxl import load_workbook

    r = client.get(
        f"{settings.API_V1_STR}/qualifications/template",
        headers=superuser_token_headers,
    )
    assert r.status_code == 200
    wb = load_workbook(BytesIO(r.content))
    notes_ws = wb["填写说明"]
    all_text = "\n".join(str(c.value) for row in notes_ws.iter_rows() for c in row if c.value)
    assert "法人证件类型" in all_text
    assert "选填" in all_text


def test_template_signature_example_has_no_brackets(
    client: TestClient, superuser_token_headers: dict[str, str]
) -> None:
    from io import BytesIO
    from openpyxl import load_workbook

    r = client.get(
        f"{settings.API_V1_STR}/qualifications/template",
        headers=superuser_token_headers,
    )
    wb = load_workbook(BytesIO(r.content))
    ws = wb.active
    sig_cell = ws.cell(row=2, column=24).value  # 短信签名列
    assert "【" not in str(sig_cell)
    assert "】" not in str(sig_cell)


def test_import_collects_all_errors_and_writes_valid_rows(
    client: TestClient, superuser_token_headers: dict[str, str]
) -> None:
    """混合文件：错误行收集所有错误，有效行批量写入"""
    headers = ["企业名称", "是否签名校验", "短信签名"]
    rows = [
        ["", "是", "签名A"],            # 行2：企业名称为空
        ["测试企业错误行", "也许", "签名B"],  # 行3：布尔字段值无效
        ["测试企业有效行", "是", "签名C"],    # 行4：有效
        ["", "", ""],                   # 行5：空行跳过
    ]
    data = _build_xlsx(headers, rows)

    files = {"file": ("test.xlsx", data, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
    r = client.post(
        f"{settings.API_V1_STR}/qualifications/import",
        headers=superuser_token_headers,
        files=files,
    )
    assert r.status_code == 200
    body = r.json()
    assert body["success_count"] == 1
    assert body["error_count"] == 2
    assert body["total"] == 3
    err_fields = {e["field"] for e in body["errors"]}
    assert err_fields == {"企业名称", "是否签名校验"}
    # 有效行确实写入
    list_r = client.get(
        f"{settings.API_V1_STR}/qualifications",
        headers=superuser_token_headers,
        params={"enterprise_name": "测试企业有效行"},
    )
    assert list_r.status_code == 200
    assert list_r.json()["total"] >= 1


def test_import_returns_error_report_xlsx(
    client: TestClient, superuser_token_headers: dict[str, str]
) -> None:
    from io import BytesIO
    from openpyxl import load_workbook

    errors = [
        {"row": 2, "field": "企业名称", "value": "", "reason": "企业名称不能为空", "suggestion": "请填写企业名称"},
        {"row": 3, "field": "是否签名校验", "value": "也许", "reason": "布尔字段值无效", "suggestion": "请填写「是」或「否」"},
    ]
    r = client.post(
        f"{settings.API_V1_STR}/qualifications/import/error-report",
        headers=superuser_token_headers,
        json={"errors": errors},
    )
    assert r.status_code == 200
    assert "spreadsheetml" in r.headers["content-type"]
    wb = load_workbook(BytesIO(r.content))
    ws = wb.active
    assert ws.title == "导入错误报告"
    assert ws.cell(row=1, column=1).value == "行号"
    assert ws.cell(row=2, column=2).value == "企业名称"
    assert ws.cell(row=3, column=5).value == "请填写「是」或「否」"



def test_preview_qualifications_import(
    client: TestClient, superuser_token_headers: dict[str, str]
) -> None:
    """预览接口返回前 5 行数据与未识别表头"""
    headers = ["企业名称", "短信签名", "未识别列X"]
    rows = [["测试企业A", "签名A", "忽略值A"]]
    data = _build_xlsx(headers, rows)

    files = {"file": ("test.xlsx", data, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
    r = client.post(
        f"{settings.API_V1_STR}/qualifications/import/preview",
        headers=superuser_token_headers,
        files=files,
    )
    assert r.status_code == 200
    body = r.json()
    assert "企业名称" in body["headers"]
    assert body["unrecognized_headers"] == ["未识别列X"]
    assert body["total_data_rows"] == 1
    assert body["rows"][0]["enterprise_name"] == "测试企业A"
    assert body["rows"][0]["sms_signature"] == "签名A"


def test_preview_qualifications_import_rejects_bad_extension(
    client: TestClient, superuser_token_headers: dict[str, str]
) -> None:
    """非 Excel 文件返回 400"""
    files = {"file": ("test.txt", b"not excel", "text/plain")}
    r = client.post(
        f"{settings.API_V1_STR}/qualifications/import/preview",
        headers=superuser_token_headers,
        files=files,
    )
    assert r.status_code == 400
