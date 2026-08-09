"""Tests for filing-tasks API: explicit port selection."""
import uuid
from collections.abc import Generator
from io import BytesIO

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlmodel import Session, delete

from app.core.config import settings
from app.core.db import engine
from app.core.storage import get_storage
from app.models import FilingTask


@pytest.fixture(scope="module", autouse=True)
def _cleanup_filing_tasks() -> Generator[None, None, None]:
    # Ensure session-scope user cleanup in conftest does not trip FK on
    # filing_task rows created by these tests.
    yield
    with Session(engine) as session:
        session.execute(delete(FilingTask))
        session.commit()


def _create_qualification(client, headers, name="测试企业"):
    r = client.post(
        f"{settings.API_V1_STR}/qualifications",
        headers=headers,
        json={
            "enterprise_name": name,
            "signature": "签名X",
        },
    )
    assert r.status_code == 200, r.text
    return r.json()["id"]


def _create_port(client, headers, main_port_number, sub_port_number=None):
    payload = {
        "carrier": "中国移动",
        "enterprise_name": "测试备案公司",
        "main_port_number": main_port_number,
        "group_code": "G001",
        "carrier_room": "机房A",
        "enterprise_room": "机房B",
        "port_type": "短信",
    }
    if sub_port_number is not None:
        payload["sub_port_number"] = sub_port_number
    r = client.post(
        f"{settings.API_V1_STR}/port-info",
        headers=headers,
        json=payload,
    )
    assert r.status_code == 200, r.text
    return r.json()["id"]


def _create_export_group(client, headers, name="导出组"):
    r = client.post(
        f"{settings.API_V1_STR}/export-groups",
        headers=headers,
        json={
            "name": name,
            "fields": [
                {"field_name": "main_port_number", "field_label": "主端口号", "sort_order": 1},
                {"field_name": "sub_port_number", "field_label": "子端口号", "sort_order": 2},
            ],
        },
    )
    assert r.status_code == 200, r.text
    return r.json()["id"]


def test_create_filing_task_with_explicit_port_ids(
    client: TestClient, superuser_token_headers: dict[str, str]
) -> None:
    qual_id = _create_qualification(client, superuser_token_headers)
    port_id_main = _create_port(client, superuser_token_headers, "10698")
    port_id_sub = _create_port(client, superuser_token_headers, "10698", "0001")
    group_id = _create_export_group(client, superuser_token_headers)

    r = client.post(
        f"{settings.API_V1_STR}/filing-tasks",
        headers=superuser_token_headers,
        json={
            "qualification_ids": [qual_id],
            "port_ids": [port_id_main, port_id_sub],
            "export_group_id": group_id,
        },
    )
    assert r.status_code == 200, r.text
    body = r.json()
    assert body["qualification_count"] == 1
    assert body["port_count"] == 2
    assert len(body["port_ids"]) == 2
    assert body["download_url"]

    # 下载并校验 Excel 包含两个端口行（资质×端口=2行 + 表头）
    r2 = client.get(
        f"{settings.API_V1_STR}/filing-tasks/{body['id']}/download",
        headers=superuser_token_headers,
    )
    assert r2.status_code == 200
    wb = load_workbook(BytesIO(r2.content))
    ws = wb.active
    assert ws.cell(row=1, column=1).value == "主端口号"
    assert ws.cell(row=1, column=2).value == "子端口号"
    # 资质(1) × 端口(2) = 2 数据行
    assert ws.max_row == 3


def test_create_filing_task_rejects_empty_port_ids(
    client: TestClient, superuser_token_headers: dict[str, str]
) -> None:
    qual_id = _create_qualification(client, superuser_token_headers)
    group_id = _create_export_group(client, superuser_token_headers)

    r = client.post(
        f"{settings.API_V1_STR}/filing-tasks",
        headers=superuser_token_headers,
        json={
            "qualification_ids": [qual_id],
            "port_ids": [],
            "export_group_id": group_id,
        },
    )
    assert r.status_code == 400
    assert "至少选择一个端口" in r.json()["detail"]


def test_create_filing_task_rejects_invalid_port_ids(
    client: TestClient, superuser_token_headers: dict[str, str]
) -> None:
    qual_id = _create_qualification(client, superuser_token_headers)
    group_id = _create_export_group(client, superuser_token_headers)
    fake_id = "00000000-0000-0000-0000-000000000000"

    r = client.post(
        f"{settings.API_V1_STR}/filing-tasks",
        headers=superuser_token_headers,
        json={
            "qualification_ids": [qual_id],
            "port_ids": [fake_id],
            "export_group_id": group_id,
        },
    )
    assert r.status_code == 400
    assert "无效" in r.json()["detail"]


def test_field_registry_uses_legal_rep_id_card_labels():
    from app.services.export_field_registry import field_map
    m = field_map()
    assert m.get("handler_id_front") == "法人身份证正面"
    assert m.get("handler_id_back") == "法人身份证反面"
    assert "经办人身份证正面" not in m.values()
    assert "经办人身份证反面" not in m.values()


def _create_task(client, headers) -> dict:
    qual_id = _create_qualification(client, headers)
    port_id = _create_port(client, headers, "10701")
    group_id = _create_export_group(client, headers)
    r = client.post(
        f"{settings.API_V1_STR}/filing-tasks",
        headers=headers,
        json={
            "qualification_ids": [qual_id],
            "port_ids": [port_id],
            "export_group_id": group_id,
        },
    )
    assert r.status_code == 200, r.text
    return r.json()


def test_download_no_file_path_returns_not_generated_reason(
    client: TestClient, superuser_token_headers: dict[str, str]
) -> None:
    task = _create_task(client, superuser_token_headers)
    with Session(engine) as session:
        db_task = session.get(FilingTask, uuid.UUID(task["id"]))
        db_task.file_path = None
        session.add(db_task)
        session.commit()

    r = client.get(
        f"{settings.API_V1_STR}/filing-tasks/{task['id']}/download",
        headers=superuser_token_headers,
    )
    assert r.status_code == 404
    detail = r.json()["detail"]
    assert detail["reason"] == "文件尚未生成"
    assert detail["can_retry"] is True


def test_download_missing_file_returns_expired_reason(
    client: TestClient, superuser_token_headers: dict[str, str]
) -> None:
    task = _create_task(client, superuser_token_headers)
    # 模拟文件过期：直接从存储删除
    storage = get_storage()
    storage.delete(task["file_path"])

    r = client.get(
        f"{settings.API_V1_STR}/filing-tasks/{task['id']}/download",
        headers=superuser_token_headers,
    )
    assert r.status_code == 404
    detail = r.json()["detail"]
    assert detail["reason"] == "文件已过期或已被删除"
    assert detail["can_retry"] is True


def test_regenerate_restores_downloadable_file(
    client: TestClient, superuser_token_headers: dict[str, str]
) -> None:
    task = _create_task(client, superuser_token_headers)
    # 模拟文件过期后下载失败
    storage = get_storage()
    storage.delete(task["file_path"])
    r = client.get(
        f"{settings.API_V1_STR}/filing-tasks/{task['id']}/download",
        headers=superuser_token_headers,
    )
    assert r.status_code == 404

    # 重新生成
    r2 = client.post(
        f"{settings.API_V1_STR}/filing-tasks/{task['id']}/regenerate",
        headers=superuser_token_headers,
    )
    assert r2.status_code == 200, r2.text
    assert r2.json()["task_id"] == task["id"]

    # 再次下载成功且内容有效
    r3 = client.get(
        f"{settings.API_V1_STR}/filing-tasks/{task['id']}/download",
        headers=superuser_token_headers,
    )
    assert r3.status_code == 200
    wb = load_workbook(BytesIO(r3.content))
    ws = wb.active
    assert ws.cell(row=1, column=1).value == "主端口号"


def test_regenerate_unknown_task_returns_404(
    client: TestClient, superuser_token_headers: dict[str, str]
) -> None:
    fake_id = "00000000-0000-0000-0000-000000000000"
    r = client.post(
        f"{settings.API_V1_STR}/filing-tasks/{fake_id}/regenerate",
        headers=superuser_token_headers,
    )
    assert r.status_code == 404
