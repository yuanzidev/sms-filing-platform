"""Tests for Excel image extraction row-index remapping (empty rows skipped in data list)."""

import io

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from PIL import Image as PILImage

from app.services.excel_image_extractor import (
    extract_cell_images_from_xlsx,
    extract_images_from_xlsx,
    inject_cell_images,
)


def _png_bytes() -> bytes:
    pil = PILImage.new("RGB", (10, 10), color="red")
    buf = io.BytesIO()
    pil.save(buf, format="PNG")
    return buf.getvalue()


def test_extract_images_from_xlsx_remaps_row_indices() -> None:
    """浮动图片：Excel 行 2/4 有数据（行 3 为空），图片应映射到数据列表索引 0/1。"""
    png = _png_bytes()
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "运营商"
    ws["B1"] = "授权书图片"
    ws["A2"] = "中国移动"
    ws["B2"] = "img"
    # 行 3 为空（数据列表中跳过）
    ws["A4"] = "中国联通"
    ws["B4"] = "img"

    ws.add_image(XLImage(io.BytesIO(png)), "B2")
    ws.add_image(XLImage(io.BytesIO(png)), "B4")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    images = extract_images_from_xlsx(buf.getvalue(), data_row_indices=[2, 4])
    assert sorted(img.row_index for img in images) == [0, 1]


def test_extract_images_from_xlsx_sets_field_name_from_headers() -> None:
    """浮动图片：按锚点列匹配表头，保存附件时能知道图片字段。"""
    png = _png_bytes()
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "运营商"
    ws["B1"] = "授权书图片"
    ws["A2"] = "中国移动"
    ws["B2"] = "img"
    ws.add_image(XLImage(io.BytesIO(png)), "B2")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    images = extract_images_from_xlsx(
        buf.getvalue(),
        headers=["运营商", "授权书图片"],
        data_row_indices=[2],
    )
    assert [img.field_name for img in images] == ["授权书图片"]


def test_extract_images_from_xlsx_drops_images_on_empty_rows() -> None:
    """浮动图片：位于空行（不在 data_row_indices 中）的图片应被移除。"""
    png = _png_bytes()
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "运营商"
    ws["B1"] = "授权书图片"
    ws["A2"] = "中国移动"
    ws["B2"] = "img"
    # 行 3 为空（数据列表中跳过），但其上挂了一张浮动图
    ws["B3"] = "img"

    ws.add_image(XLImage(io.BytesIO(png)), "B2")
    ws.add_image(XLImage(io.BytesIO(png)), "B3")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    images = extract_images_from_xlsx(buf.getvalue(), data_row_indices=[2])
    assert [img.row_index for img in images] == [0]


def test_extract_cell_images_from_xlsx_remaps_row_indices() -> None:
    """单元格内嵌图片：Excel 行 2/4 有数据（行 3 为空），图片应映射到数据列表索引 0/1。"""
    png = _png_bytes()
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "运营商"
    ws["B1"] = "授权书图片"
    ws["A2"] = "中国移动"
    ws["B2"] = "img"
    # 行 3 为空（数据列表中跳过）
    ws["A4"] = "中国联通"
    ws["B4"] = "img"

    buf = io.BytesIO()
    wb.save(buf)
    xlsx = inject_cell_images(buf.getvalue(), {"B2": png, "B4": png})

    images = extract_cell_images_from_xlsx(
        xlsx, headers=["运营商", "授权书图片"], data_row_indices=[2, 4]
    )
    assert sorted(img.row_index for img in images) == [0, 1]
    assert {img.field_name for img in images} == {"授权书图片"}
