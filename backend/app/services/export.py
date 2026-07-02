"""Excel export service — generates .xlsx from filing records using export group config."""
import io
import uuid
from datetime import date, datetime
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from sqlmodel import Session

from app.crud.export_group import get_export_group
from app.models import FilingRecord, PortInfo, QualificationInfo


def build_field_map() -> dict[str, str]:
    """Map logical field_name to the model attribute path used in export.
    Returns {field_name: label} for all supported fields."""
    return {
        # Port info fields
        "carrier": "运营商",
        "operation_type": "操作类型",
        "main_port_number": "主端口号",
        "sub_port_number": "子端口号",
        "port_range": "码号使用范围",
        "province": "接入省",
        "city": "接入地市",
        "port_type": "端口类型",
        "port_activation_date": "端口入网时间",
        "allow_self_extension": "是否允许自行扩展",
        "business_attribute": "业务属性",
        "business_type": "业务类型",
        "business_subtype": "业务细类",
        "specific_usage": "具体用途",
        "sms_signature": "短信签名",
        "is_gateway_signature": "是否网关签名",
        "carrier_room": "运营商接入机房及设备",
        "enterprise_room": "企业接入机房及设备",
        "has_authorization": "是否具有授权书",
        "auth_start_date": "授权开始日期",
        "auth_end_date": "授权结束日期",
        "sms_template_content": "短信模板内容",
        # Qualification info fields
        "submit_unit": "报送单位",
        "carrier_enterprise_id": "运营商企业ID",
        "enterprise_name": "企业名称",
        "cert_type": "单位证件类型",
        "cert_number": "单位证件号码",
        "app_platform_name": "APP/平台名称",
        "group_code": "集团编码",
        "responsible_name": "责任人姓名",
        "responsible_cert_type": "责任人证件类型",
        "responsible_cert_number": "责任人证件号码",
        "responsible_phone": "责任人手机号",
        "handler_name": "经办人姓名",
        "handler_cert_type": "经办人证件类型",
        "handler_cert_number": "经办人证件号码",
        "handler_phone": "经办人手机号",
        # Record info fields
        "record_number": "报备编号",
        "status": "状态",
        "created_at": "创建时间",
    }


def get_field_value(obj: FilingRecord, field_name: str) -> str:
    """Get field value from port_info, qualification_info, or filing_record."""
    pi_fields = {
        "carrier", "operation_type", "main_port_number", "sub_port_number",
        "port_range", "province", "city", "port_type", "port_activation_date",
        "allow_self_extension", "business_attribute", "business_type",
        "business_subtype", "specific_usage", "sms_signature",
        "is_gateway_signature", "carrier_room", "enterprise_room",
        "has_authorization", "auth_start_date", "auth_end_date",
        "sms_template_content",
    }
    qi_fields = {
        "submit_unit", "carrier_enterprise_id", "enterprise_name", "cert_type",
        "cert_number", "app_platform_name", "group_code", "responsible_name",
        "responsible_cert_type", "responsible_cert_number", "responsible_phone",
        "handler_name", "handler_cert_type", "handler_cert_number", "handler_phone",
    }

    if field_name in pi_fields and obj.port_info:
        value = getattr(obj.port_info, field_name, "")
    elif field_name in qi_fields and obj.qualification_info:
        value = getattr(obj.qualification_info, field_name, "")
    else:
        value = getattr(obj, field_name, "")

    if value is None:
        return ""
    if isinstance(value, bool):
        return "是" if value else "否"
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    return str(value)


def generate_export(
    session: Session,
    export_group_id: uuid.UUID,
    filters: dict[str, Any] | None = None,
) -> io.BytesIO:
    """Generate an Excel file based on export group field configuration."""
    group = get_export_group(session=session, id=export_group_id)
    if not group:
        raise ValueError("Export group not found")

    # Get records matching filters
    from app.crud.record import list_filing_records

    records, _ = list_filing_records(
        session=session, skip=0, limit=100000, **(filters or {})
    )

    field_map = build_field_map()
    wb = Workbook()
    ws = wb.active
    ws.title = "报备记录导出"

    # Header style
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(
        start_color="4472C4", end_color="4472C4", fill_type="solid"
    )
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    # Sort fields by sort_order and filter to known field names
    sorted_fields = sorted(group.fields, key=lambda f: f.sort_order)
    col_names = [f.field_name for f in sorted_fields if f.field_name in field_map]

    # Write headers
    for col_idx, field_name in enumerate(col_names, 1):
        cell = ws.cell(row=1, column=col_idx, value=field_map[field_name])
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Write data rows
    for row_idx, record in enumerate(records, 2):
        for col_idx, field_name in enumerate(col_names, 1):
            value = get_field_value(record, field_name)
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border

    # Auto-adjust column widths
    for col_idx, _ in enumerate(col_names, 1):
        max_width = 0
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx, values_only=True):
            for cell_value in row:
                if cell_value:
                    max_width = max(max_width, len(str(cell_value)))
        letter = ws.cell(row=1, column=col_idx).column_letter
        ws.column_dimensions[letter].width = min(max_width + 4, 50)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output
