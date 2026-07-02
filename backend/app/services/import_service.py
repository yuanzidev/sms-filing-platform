"""Excel import service with embedded image extraction."""
import hashlib
import io
import uuid
from datetime import date
from pathlib import Path
from typing import Any

import openpyxl
from sqlmodel import Session

from app.core.config import settings
from app.core.storage import get_storage


def parse_excel_preview(file_content: bytes, filename: str) -> dict[str, Any]:
    """Step 1: Parse uploaded Excel and return headers + preview + image info.

    Returns dict with:
      - headers: list[str] — column headers from row 1
      - preview_rows: list[list] — first 20 data rows (cell values)
      - image_columns: dict[int, str] — {col_index: "column_name"} for columns with images
      - total_rows: int — total data rows (excluding header)
      - image_count: int — total embedded images found
      - file_token: str — token to reference the temp file for confirm step
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_content), data_only=True)
    ws = wb.active

    # Extract headers (row 1)
    headers = [str(cell.value) if cell.value is not None else "" for cell in next(ws.iter_rows(min_row=1, max_row=1))]

    # Build (row, col) -> image mapping
    image_map: dict[tuple[int, int], list[bytes]] = {}
    image_count = 0

    try:
        for image in ws._images:
            row = image.anchor._from.row + 1  # 1-based
            col = image.anchor._from.col + 1
            img_data = image._data()
            image_map.setdefault((row, col), []).append(img_data)
            image_count += 1
    except Exception:
        pass

    # Detect which columns have images in the first 20 rows
    image_columns: dict[int, str] = {}
    for (row, col), imgs in image_map.items():
        if row <= 21 and col <= len(headers):
            image_columns[col] = headers[col - 1] if col <= len(headers) else f"列{col}"

    # Read preview rows (rows 2-21)
    preview_rows: list[list] = []
    total_rows = ws.max_row - 1  # Exclude header
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=min(21, ws.max_row), values_only=True), start=2):
        cells = []
        for col_idx, cell_value in enumerate(row, start=1):
            has_image = (row_idx, col_idx) in image_map
            cells.append({
                "value": str(cell_value) if cell_value is not None else "",
                "has_image": has_image,
            })
        preview_rows.append(cells)

    # Save temp file for confirm step
    temp_dir = Path(settings.LOCAL_STORAGE_DIR) / "imports"
    temp_dir.mkdir(parents=True, exist_ok=True)
    file_token = uuid.uuid4().hex
    temp_path = temp_dir / f"{file_token}.xlsx"
    temp_path.write_bytes(file_content)

    wb.close()
    return {
        "headers": headers,
        "preview_rows": preview_rows,
        "image_columns": image_columns,
        "total_rows": total_rows,
        "image_count": image_count,
        "file_token": file_token,
    }


def confirm_import(
    session: Session,
    file_token: str,
    field_mapping: dict[str, str],
    operator_id: uuid.UUID | None,
) -> dict[str, Any]:
    """Step 2: Confirm import with field mapping. Process rows in batches.

    field_mapping: {column_index: "field_name"} — maps Excel column to target field name

    Returns: {success_count, error_count, errors: [{row, message}]}
    """
    temp_dir = Path(settings.LOCAL_STORAGE_DIR) / "imports"
    temp_path = temp_dir / f"{file_token}.xlsx"
    if not temp_path.exists():
        raise FileNotFoundError("Import file expired, please re-upload")

    wb = openpyxl.load_workbook(str(temp_path))
    ws = wb.active

    # Build image map
    image_map: dict[tuple[int, int], list[bytes]] = {}
    try:
        for image in ws._images:
            row = image.anchor._from.row + 1
            col = image.anchor._from.col + 1
            image_map.setdefault((row, col), []).append(image._data())
    except Exception:
        pass

    # Parse field_mapping keys to int (Excel column index -> field name)
    col_to_field: dict[int, str] = {}
    pi_fields = {
        "carrier", "operation_type", "main_port_number", "sub_port_number", "port_range",
        "province", "city", "port_type", "port_activation_date", "allow_self_extension",
        "business_attribute", "business_type", "business_subtype", "specific_usage",
        "sms_signature", "is_gateway_signature", "carrier_room", "enterprise_room",
        "has_authorization", "auth_start_date", "auth_end_date", "sms_template_content",
    }
    qi_fields = {
        "submit_unit", "carrier_enterprise_id", "enterprise_name", "cert_type",
        "cert_number", "app_platform_name", "group_code", "responsible_name",
        "responsible_cert_type", "responsible_cert_number", "responsible_phone",
        "handler_name", "handler_cert_type", "handler_cert_number", "handler_phone",
    }

    for col_str, field_name in field_mapping.items():
        col_to_field[int(col_str)] = field_name

    import_batch = f"IMP-{date.today().isoformat()}-{uuid.uuid4().hex[:8]}"
    storage = get_storage()
    success_count = 0
    errors: list[dict] = []

    from app.models import FileAttachment, FilingRecord, PortInfo, QualificationInfo

    rows = list(ws.iter_rows(min_row=2, values_only=True))
    batch_size = 200

    for batch_start in range(0, len(rows), batch_size):
        batch_rows = rows[batch_start:batch_start + batch_size]
        for row_offset, row in enumerate(batch_rows):
            row_idx = batch_start + row_offset + 2  # 1-based, header is row 1
            try:
                pi_data: dict[str, Any] = {}
                qi_data: dict[str, Any] = {}
                image_attachments: list[tuple[int, bytes]] = []  # (col_idx, img_bytes)

                for col_idx, field_name in col_to_field.items():
                    cell_value = row[col_idx - 1] if col_idx <= len(row) else None

                    # Check for images at this cell
                    if (row_idx, col_idx) in image_map:
                        for img_data in image_map[(row_idx, col_idx)]:
                            image_attachments.append((col_idx, img_data))

                    value = str(cell_value).strip() if cell_value is not None else ""

                    if field_name in pi_fields:
                        pi_data[field_name] = value
                    elif field_name in qi_fields:
                        qi_data[field_name] = value

                # Create port_info
                pi = PortInfo(**pi_data)
                session.add(pi)
                session.flush()

                # Create qualification_info
                qi = QualificationInfo(**qi_data)
                session.add(qi)
                session.flush()

                # Generate record number
                from app.crud.record import _record_number_sequence
                seq = _record_number_sequence(session)
                record_number = f"REC-{date.today().strftime('%Y%m%d')}-{seq:04d}"

                fr = FilingRecord(
                    record_number=record_number,
                    status="已报备",
                    port_info_id=pi.id,
                    qualification_info_id=qi.id,
                    operator_id=operator_id,
                    import_batch=import_batch,
                    source_file=temp_path.name,
                )
                session.add(fr)
                session.flush()

                # Upload images to storage and create file_attachment records
                for img_col, img_data in image_attachments:
                    md5_hash = hashlib.md5(img_data).hexdigest()
                    ext = ".png"
                    key = f"images/{date.today().isoformat()}/{uuid.uuid4().hex}{ext}"
                    storage.upload(key, img_data, "image/png")

                    attachment = FileAttachment(
                        original_name=f"cell_{row_idx}_{img_col}{ext}",
                        stored_path=key,
                        file_size=len(img_data),
                        mime_type="image/png",
                        md5_hash=md5_hash,
                        entity_type="port_info",
                        entity_id=pi.id,
                        uploader_id=operator_id,
                    )
                    session.add(attachment)

                success_count += 1

            except Exception as e:
                session.rollback()
                errors.append({"row": row_idx, "message": str(e)})

        session.commit()

    wb.close()

    # Clean up temp file
    try:
        temp_path.unlink()
    except Exception:
        pass

    return {
        "success_count": success_count,
        "error_count": len(errors),
        "errors": errors,
        "import_batch": import_batch,
    }
