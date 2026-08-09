"""Qualification info management routes."""
import io
import uuid
from typing import Any
from urllib.parse import quote

from fastapi import APIRouter, Depends, File, HTTPException, Query, Request, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from pydantic import BaseModel

from app.api.deps import CurrentUser, SessionDep, get_current_active_superuser
from app.crud.qualification import (
    create_qualification,
    delete_qualification,
    get_qualification,
    get_qualifications_by_signatures,
    list_qualifications,
    update_qualification,
)
from app.models import (
    BatchSignatureRequest,
    BatchSignatureResponse,
    Message,
    QualificationInfo,
    QualificationInfoCreate,
    QualificationInfoPublic,
    QualificationInfosPublic,
    QualificationInfoUpdate,
)
from app.services.excel_image_extractor import (
    extract_cell_images_from_xlsx,
    extract_images_from_xlsx,
    upload_import_images,
)
from app.services.operation_log import log_operation

router = APIRouter(
    prefix="/qualifications",
    tags=["qualifications"],
    dependencies=[Depends(get_current_active_superuser)],
)

_QUALIFICATION_HEADERS = [
    "企业名称",
    "单位证件号码",
    "单位证件类型",
    "单位证件图片",
    "责任人姓名",
    "责任人证件类型",
    "责任人证件号码",
    "责任人手机号",
    "责任人证件地址",
    "责任人身份证正面",
    "责任人身份证反面",
    "经办人姓名",
    "经办人证件类型",
    "经办人证件号码",
    "经办人手机号",
    "经办人证件地址",
    "经办人现场照片",
    "法人姓名",
    "法人证件类型",
    "法人证件号码",
    "法人证件地址",
    "法人身份证正面",
    "法人身份证反面",
    "短信签名",
    "签名类型/来源",
    "短信模板内容",
    "是否签名校验",
    "是否网关签名",
    "模板是否包含变量",
    "模板参数类型",
    "模板参数长度",
    "业务属性",
    "业务类型",
    "业务细类",
    "具体用途",
    "引流号码",
    "引流链接",
    "引流号码类型",
    "引流号码用途",
    "引流内容",
    "链接类型",
    "签名举证附件",
    "引流号码举证附件",
    "引流链接举证",
    "APP/平台名称",
]


@router.get("/template")
def download_qualification_template() -> Any:
    from openpyxl.styles import Font
    from PIL import Image, ImageDraw
    from app.services.excel_image_extractor import inject_cell_images

    wb = Workbook()
    ws = wb.active
    ws.title = "资质导入模板"

    for col_idx, header in enumerate(_QUALIFICATION_HEADERS, 1):
        ws.cell(row=1, column=col_idx, value=header)

    example_data = [
        "示例企业有限公司",
        "91110108MA01XXXXX",          # 单位证件号码
        "营业执照",                    # 单位证件类型
        "",                            # 单位证件图片
        "李四",                        # 责任人姓名
        "身份证",                      # 责任人证件类型
        "110101199001011234",          # 责任人证件号码
        "13800138000",                 # 责任人手机号
        "北京市朝阳区XX路1号",         # 责任人证件地址
        "",                            # 责任人身份证正面
        "",                            # 责任人身份证反面
        "王五",                        # 经办人姓名
        "身份证",                      # 经办人证件类型
        "110101199501011234",          # 经办人证件号码
        "13900139000",                 # 经办人手机号
        "北京市海淀区XX路2号",         # 经办人证件地址
        "",                            # 经办人现场照片
        "张三",                        # 法人姓名
        "身份证",                      # 法人证件类型
        "110101199001011234",          # 法人证件号码
        "北京市朝阳区XX路1号",         # 法人证件地址
        "",                            # 法人身份证正面
        "",                            # 法人身份证反面
        "示例平台",                    # 短信签名
        "自营签名",                    # 签名类型/来源
        "您的验证码是{code}，请在5分钟内完成验证",  # 短信模板内容
        "是",                          # 是否签名校验
        "否",                          # 是否网关签名
        "是",                          # 模板是否包含变量
        "数字",                        # 模板参数类型
        "6",                           # 模板参数长度
        "营销类",                      # 业务属性
        "验证码",                      # 业务类型
        "登录验证",                    # 业务细类
        "用户登录验证",                # 具体用途
        "13800000000",                 # 引流号码
        "https://example.com",         # 引流链接
        "手机号",                      # 引流号码类型
        "业务联系",                    # 引流号码用途
        "欢迎使用我们的服务",          # 引流内容
        "H5",                          # 链接类型
        "",                            # 签名举证附件
        "",                            # 引流号码举证附件
        "",                            # 引流链接举证
        "示例平台",                    # APP/平台名称
    ]
    for col_idx, val in enumerate(example_data, 1):
        ws.cell(row=2, column=col_idx, value=val)

    instructions = wb.create_sheet("填写说明")
    instructions.cell(row=1, column=1, value="Excel 导入图片填写说明").font = Font(bold=True, size=14)
    notes = [
        "1. 请勿修改表头行（第一行）的列标题",
        "2. 每条数据填写一行，从第二行开始",
        "3. 图片列（单位证件图片、身份证正面/反面等）用于存放资质证明图片",
        "4. 插入方法：右键单元格 ->「插入图片」->「放置在单元格中」-> 选择图片文件",
        "5. 也可将图片直接拖入到图片列的单元格中",
        "6. 系统会自动提取每行单元格内嵌的图片，并与对应字段关联",
        "7. 支持的图片格式：PNG、JPEG、GIF、BMP、WEBP，单张不超过 10MB",
        "8. 法人证件类型/号码/地址：选填；运营商报备强依赖时再填",
        "9. 支持图片的列：单位证件图片、责任人身份证正面/反面、法人身份证正面/反面、签名举证附件、引流号码举证附件、引流链接举证、经办人现场照片；图片文件建议小于 10MB，支持 PNG、JPEG 格式",
    ]
    for i, note in enumerate(notes, 2):
        instructions.cell(row=i, column=1, value=note)

    output = io.BytesIO()
    wb.save(output)
    xlsx_bytes = output.getvalue()

    # Inject sample cell images into the template
    sample_img = Image.new("RGB", (120, 60), color=(220, 230, 241))
    draw = ImageDraw.Draw(sample_img)
    draw.text((10, 20), "示例\n请替换", fill=(50, 50, 50))
    img_buf = io.BytesIO()
    sample_img.save(img_buf, format="PNG")

    # 签名举证附件 is column 42 (1-based) = "AP2"
    cell_images = {"AP2": img_buf.getvalue()}
    xlsx_bytes = inject_cell_images(xlsx_bytes, cell_images)

    return StreamingResponse(
        io.BytesIO(xlsx_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote('资质导入模板.xlsx')}"},
    )


@router.post("/import")
def import_qualifications(
    *, session: SessionDep, file: UploadFile = File(...)
) -> Any:
    if not file.filename or not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="仅支持 .xlsx 或 .xls 文件")

    content = file.file.read()
    try:
        wb = load_workbook(io.BytesIO(content))
    except Exception:
        raise HTTPException(status_code=400, detail="无法解析 Excel 文件，请检查文件格式")

    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 1:
        raise HTTPException(status_code=400, detail="文件为空，请导入有效的 Excel 文件")

    # Map model field names to column index by matching Chinese headers
    header_to_field = {
        "企业名称": "enterprise_name",
        "单位证件类型": "cert_type",
        "单位证件号码": "cert_number",
        "APP/平台名称": "app_platform_name",
        "法人姓名": "legal_representative_name",
        "法人证件类型": "legal_representative_cert_type",
        "法人证件号码": "legal_representative_cert_number",
        "法人证件地址": "legal_representative_cert_address",
        "责任人姓名": "responsible_name",
        "责任人证件类型": "responsible_cert_type",
        "责任人证件号码": "responsible_cert_number",
        "责任人证件地址": "responsible_address",
        "责任人手机号": "responsible_phone",
        "经办人姓名": "handler_name",
        "经办人证件类型": "handler_cert_type",
        "经办人证件号码": "handler_cert_number",
        "经办人证件地址": "handler_address",
        "经办人手机号": "handler_phone",
        "短信签名": "sms_signature",
        "签名类型/来源": "signature_type",
        "是否签名校验": "signature_verified",
        "是否网关签名": "is_gateway_signature",
        "短信模板内容": "sms_template_content",
        "模板是否包含变量": "template_has_variable",
        "模板参数类型": "template_param_type",
        "模板参数长度": "template_param_length",
        "业务属性": "business_attribute",
        "业务类型": "business_type",
        "业务细类": "business_subtype",
        "具体用途": "specific_usage",
        "引流号码": "diversion_number",
        "引流号码类型": "diversion_number_type",
        "引流号码用途": "diversion_number_usage",
        "引流内容": "diversion_content",
        "引流链接": "link_address",
        "链接类型": "link_type",
    }

    header_row = [str(c) if c else "" for c in rows[0]]
    col_map: dict[str, int] = {}
    for col_idx, h in enumerate(header_row):
        if h in header_to_field:
            col_map[header_to_field[h]] = col_idx

    missing = [h for h, f in header_to_field.items() if f == "enterprise_name" and f not in col_map]
    if missing:
        raise HTTPException(
            status_code=400,
            detail=f"模板不匹配：缺少必填列「{'」「'.join(missing)}」。请确认使用了正确的资质导入模板（首页 → 资质管理 → 下载模板）",
        )

    # Phase 1: Validate all rows, collect errors
    objects: list[QualificationInfo] = []
    errors: list[dict] = []
    data_row_indices: list[int] = []

    for row_idx, row in enumerate(rows[1:], start=2):
        # Skip completely empty rows
        if all(c is None or str(c).strip() == "" for c in row):
            continue

        def cell(col_name: str) -> str | None:
            if col_name not in col_map:
                return None
            idx = col_map[col_name]
            if idx >= len(row):
                return None
            v = row[idx]
            if v is None or str(v).strip() == "":
                return None
            return str(v).strip()

        def parse_bool(col_name: str) -> bool | None:
            v = cell(col_name)
            if v is None:
                return None
            return v in ("是", "true", "True", "1", "TRUE")

        row_errors: list[dict] = []
        enterprise_name = cell("enterprise_name")
        if not enterprise_name:
            row_errors.append({
                "row": row_idx, "field": "企业名称", "value": "",
                "reason": "企业名称不能为空", "suggestion": "请填写企业名称",
            })

        # Validate booleans
        for bool_field, cn_name in [
            ("signature_verified", "是否签名校验"),
            ("is_gateway_signature", "是否网关签名"),
            ("template_has_variable", "模板是否包含变量"),
        ]:
            v = cell(bool_field)
            if v and v not in ("是", "否", "true", "True", "1", "TRUE", "false", "False", "0", "FALSE"):
                row_errors.append({
                    "row": row_idx, "field": cn_name, "value": v,
                    "reason": "布尔字段值无效", "suggestion": "请填写「是」或「否」",
                })

        if row_errors:
            errors.extend(row_errors)
        else:
            legal_rep_cert_type = cell("legal_representative_cert_type")
            legal_rep_cert_number = cell("legal_representative_cert_number")
            legal_rep_cert_address = cell("legal_representative_cert_address")

            objects.append(QualificationInfo(
                enterprise_name=enterprise_name,
                cert_type=cell("cert_type"),
                cert_number=cell("cert_number"),
                app_platform_name=cell("app_platform_name"),
                legal_representative_name=cell("legal_representative_name"),
                legal_representative_cert_type=legal_rep_cert_type,
                legal_representative_cert_number=legal_rep_cert_number,
                legal_representative_cert_address=legal_rep_cert_address,
                responsible_name=cell("responsible_name"),
                responsible_cert_type=cell("responsible_cert_type"),
                responsible_cert_number=cell("responsible_cert_number"),
                responsible_address=cell("responsible_address"),
                responsible_phone=cell("responsible_phone"),
                handler_name=cell("handler_name"),
                handler_cert_type=cell("handler_cert_type"),
                handler_cert_number=cell("handler_cert_number"),
                handler_address=cell("handler_address"),
                handler_phone=cell("handler_phone"),
                sms_signature=cell("sms_signature"),
                signature_type=cell("signature_type"),
                signature_verified=parse_bool("signature_verified"),
                is_gateway_signature=parse_bool("is_gateway_signature"),
                sms_template_content=cell("sms_template_content"),
                template_has_variable=parse_bool("template_has_variable"),
                template_param_type=cell("template_param_type"),
                template_param_length=cell("template_param_length"),
                business_attribute=cell("business_attribute"),
                business_type=cell("business_type"),
                business_subtype=cell("business_subtype"),
                specific_usage=cell("specific_usage"),
                diversion_number=cell("diversion_number"),
                diversion_number_type=cell("diversion_number_type"),
                diversion_number_usage=cell("diversion_number_usage"),
                diversion_content=cell("diversion_content"),
                link_address=cell("link_address"),
                link_type=cell("link_type"),
            ))
            data_row_indices.append(row_idx)

    # Phase 2: If no valid rows, return all errors
    if not objects and errors:
        return {
            "total": len(objects) + len({e["row"] for e in errors}),
            "success_count": 0,
            "error_count": len(errors),
            "errors": errors,
        }

    if not objects:
        raise HTTPException(status_code=400, detail="文件中没有有效数据")

    # Phase 3: Write valid rows + extract images with fixed indices
    session.add_all(objects)
    session.flush()

    warnings: list[str] = []
    if file.filename.endswith(".xlsx"):
        all_images: list = []
        try:
            all_images.extend(extract_cell_images_from_xlsx(
                content, headers=header_row, data_row_indices=data_row_indices,
            ))
        except Exception as e:
            warnings.append(f"单元格图片提取失败: {e}")
        try:
            all_images.extend(extract_images_from_xlsx(content, data_row_indices=data_row_indices))
        except Exception as e:
            warnings.append(f"浮动图片提取失败: {e}")
        if all_images:
            _, img_warnings = upload_import_images(
                images=all_images,
                objects=objects,
                entity_type="qualification_info",
                session=session,
            )
            warnings.extend(img_warnings)

    session.commit()

    return {
        "total": len(objects) + len({e["row"] for e in errors}),
        "success_count": len(objects),
        "error_count": len(errors),
        "errors": errors,
        "warnings": warnings,
    }


class ImportErrorReport(BaseModel):
    errors: list[dict]


@router.post("/import/error-report")
def download_import_error_report(body: ImportErrorReport) -> Any:
    """Generate an Excel file highlighting import errors."""
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "导入错误报告"

    headers = ["行号", "字段", "原值", "失败原因", "修复建议"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)

    red_fill = PatternFill(start_color="FFD7D7", end_color="FFD7D7", fill_type="solid")
    for i, err in enumerate(body.errors, 2):
        ws.cell(row=i, column=1, value=err.get("row"))
        ws.cell(row=i, column=2, value=err.get("field"))
        ws.cell(row=i, column=3, value=err.get("value"))
        ws.cell(row=i, column=4, value=err.get("reason"))
        ws.cell(row=i, column=5, value=err.get("suggestion"))
        for col in range(1, 6):
            ws.cell(row=i, column=col).fill = red_fill

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote('导入错误报告.xlsx')}"},
    )


@router.post("/batch-by-signatures", response_model=BatchSignatureResponse)
def batch_by_signatures(
    *, session: SessionDep, body: BatchSignatureRequest
) -> Any:
    qualified, unmatched = get_qualifications_by_signatures(
        session=session, signatures=body.signatures
    )
    return BatchSignatureResponse(
        matched_qualifications=qualified,
        unmatched_signatures=unmatched,
    )


@router.get("", response_model=QualificationInfosPublic)
@router.get("/", include_in_schema=False, response_model=QualificationInfosPublic)
def read_qualifications(
    session: SessionDep,
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    enterprise_name: str | None = None,
    cert_number: str | None = None,
    sms_signature: str | None = None,
) -> Any:
    skip = (page - 1) * page_size
    items, total = list_qualifications(
        session=session, skip=skip, limit=page_size,
        enterprise_name=enterprise_name, cert_number=cert_number, sms_signature=sms_signature,
    )
    return QualificationInfosPublic(data=items, total=total, page=page, page_size=page_size)


@router.post("", response_model=QualificationInfoPublic)
@router.post("/", include_in_schema=False, response_model=QualificationInfoPublic)
def create_qualification_endpoint(*, session: SessionDep, create: QualificationInfoCreate, current_user: CurrentUser, request: Request) -> Any:
    result = create_qualification(session=session, create=create)
    log_operation(session=session, user=current_user, user_ip=request.client.host if request.client else "", module="qualifications", action="create", target=result.enterprise_name or str(result.id))
    return result


@router.get("/{id}", response_model=QualificationInfoPublic)
def read_qualification(*, session: SessionDep, id: uuid.UUID) -> Any:
    db_obj = get_qualification(session=session, id=id)
    if not db_obj:
        raise HTTPException(status_code=404, detail="资质信息不存在")
    return db_obj


@router.patch("/{id}", response_model=QualificationInfoPublic)
def update_qualification_endpoint(
    *, session: SessionDep, id: uuid.UUID, update: QualificationInfoUpdate, current_user: CurrentUser, request: Request
) -> Any:
    db_obj = get_qualification(session=session, id=id)
    if not db_obj:
        raise HTTPException(status_code=404, detail="资质信息不存在")
    result = update_qualification(session=session, db_obj=db_obj, update=update)
    log_operation(session=session, user=current_user, user_ip=request.client.host if request.client else "", module="qualifications", action="update", target=result.enterprise_name or str(id))
    return result


@router.delete("/{id}")
def delete_qualification_endpoint(*, session: SessionDep, id: uuid.UUID, current_user: CurrentUser, request: Request) -> Message:
    db_obj = get_qualification(session=session, id=id)
    if not db_obj:
        raise HTTPException(status_code=404, detail="资质信息不存在")
    target = db_obj.enterprise_name or str(id)
    delete_qualification(session=session, db_obj=db_obj)
    log_operation(session=session, user=current_user, user_ip=request.client.host if request.client else "", module="qualifications", action="delete", target=target)
    return Message(message="资质信息删除成功")
