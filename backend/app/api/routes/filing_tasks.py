"""Filing tasks API routes — export task management."""
import io
import uuid
from datetime import date, datetime
from typing import Any

from fastapi import APIRouter, HTTPException, Query, Request
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from sqlmodel import select

from app.api.deps import CurrentUser, SessionDep
from app.core.storage import get_storage
from app.crud.export_group import get_export_group
from app.crud.filing_sub_port_usage import delete_usages_by_task
from app.crud.filing_task import (
    create_filing_task as crud_create_filing_task,
)
from app.crud.filing_task import (
    delete_filing_task,
    get_filing_task,
    list_filing_tasks,
)
from app.models import (
    ExportGroup,
    FilingTaskCreate,
    FilingTaskDetail,
    FilingTaskPublic,
    FilingTasksPublic,
    Message,
    PortInfo,
    QualificationInfo,
    User,
)
from app.services.export_field_registry import REGISTRY, field_map, field_source
from app.services.operation_log import log_operation
from app.services.sub_port_allocator import (
    MAX_RANGE_SIZE,
    AllocationMode,
    allocate_sub_ports,
)

router = APIRouter(prefix="/filing-tasks", tags=["filing-tasks"])


def get_field_value(
    qualification: QualificationInfo,
    port: PortInfo,
    field_name: str,
    allocated_sub_port: str | None = None,
) -> str:
    """Get field value via registry source dispatch."""
    if field_name == "sub_port_number" and allocated_sub_port is not None:
        return allocated_sub_port
    if field_name == "port_main_number":
        return getattr(port, "main_port_number", "") or ""
    if field_name == "port_sub_extension":
        if allocated_sub_port is not None:
            return allocated_sub_port
        return getattr(port, "sub_port_number", "") or ""
    if field_name == "port_full_number":
        mpn = getattr(port, "main_port_number", "") or ""
        sub = (
            allocated_sub_port
            if allocated_sub_port is not None
            else getattr(port, "sub_port_number", "") or ""
        )
        return mpn + sub if mpn and sub else mpn or sub or ""
    source = field_source(field_name)
    if source is None:
        return ""
    if source in ("image_qualification", "image_port"):
        return "[图片]"
    # 端口字段名 → 模型属性名 别名映射（如 port_enterprise_name → enterprise_name）
    _PORT_ALIAS: dict[str, str] = {
        "port_enterprise_name": "enterprise_name",
    }
    if source == "port":
        attr = _PORT_ALIAS.get(field_name, field_name)
        value = getattr(port, attr, "")
    elif source == "qualification":
        value = getattr(qualification, field_name, "")
    else:
        return ""

    if value is None:
        return ""
    if isinstance(value, bool):
        return "是" if value else "否"
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    return str(value)


def generate_excel(
    qualifications: list[QualificationInfo],
    ports: list[PortInfo],
    export_group: ExportGroup,
    group_by_field: str | None = None,
    qual_images: dict[uuid.UUID, dict[str, bytes]] | None = None,
    allocated_sub_ports: dict[tuple[uuid.UUID, str], str] | None = None,
    auto_allocate_sub_ports: bool = False,
) -> io.BytesIO:
    """Generate Excel bytes from qualification+port Cartesian product.

    qual_images: {qual_id: {field_name: image_bytes}} for cell-embedded images.
    field_name is the logical name (e.g. "cert_image").

    auto_allocate_sub_ports: 为 True 时只输出主端口行，子端口号取 allocated_sub_ports
    中按 (qualification_id, main_port_number) 分配的结果。
    """
    qual_images = qual_images or {}
    allocated_sub_ports = allocated_sub_ports or {}
    fm = field_map()
    img_field_names = {f.name for f in REGISTRY if f.source.startswith("image")}
    wb = Workbook()
    ws = wb.active
    ws.title = "报备导出"

    # Header style
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    sorted_fields = sorted(export_group.fields, key=lambda f: f.sort_order)
    col_names = [f.field_name for f in sorted_fields if f.field_name in fm]

    # Build column index for image fields
    img_col_map: dict[str, int] = {}
    for col_idx, field_name in enumerate(col_names, 1):
        if field_name in img_field_names:
            img_col_map[field_name] = col_idx

    # Write headers
    for col_idx, field_name in enumerate(col_names, 1):
        cell = ws.cell(row=1, column=col_idx, value=fm[field_name])
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Build rows: Cartesian product (auto 模式下仅主端口行，附带分配的子端口号)
    rows: list[tuple[QualificationInfo, PortInfo, str | None]]
    if auto_allocate_sub_ports:
        main_port_dict: dict[str, PortInfo] = {}
        for p in ports:
            if not p.sub_port_number and p.main_port_number not in main_port_dict:
                main_port_dict[p.main_port_number] = p
        rows = [
            (q, main_port_dict[mpn], allocated_sub_ports.get((q.id, mpn), ""))
            for q in qualifications
            for mpn in main_port_dict
        ]
    else:
        rows = [(q, p, None) for q in qualifications for p in ports]

    # Sort by group_by_field if specified
    if group_by_field:
        rows.sort(key=lambda r: get_field_value(r[0], r[1], group_by_field, r[2]))

    # Write data rows, inserting blank separator between groups
    row_idx = 2
    prev_group_value: str | None = None
    cell_images: dict[str, bytes] = {}

    for q, p, allocated_sub in rows:
        if group_by_field:
            current_value = get_field_value(q, p, group_by_field, allocated_sub)
            if prev_group_value is not None and current_value != prev_group_value:
                for col_idx in range(1, len(col_names) + 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value="")
                    cell.border = thin_border
                row_idx += 1
            prev_group_value = current_value

        for col_idx, field_name in enumerate(col_names, 1):
            value = get_field_value(q, p, field_name, allocated_sub)
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border

        # Collect images for this row from the qualification
        if img_col_map and q.id in qual_images:
            qual_imgs = qual_images[q.id]
            for field_name, col_idx in img_col_map.items():
                if field_name in qual_imgs:
                    col_letter = ws.cell(row=row_idx, column=col_idx).column_letter
                    cell_ref = f"{col_letter}{row_idx}"
                    cell_images[cell_ref] = qual_imgs[field_name]

        row_idx += 1

    # Auto-fit column widths
    for col_idx in range(1, len(col_names) + 1):
        max_width = 0
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx, values_only=True):
            for cell_value in row:
                if cell_value:
                    max_width = max(max_width, len(str(cell_value)))
        letter = ws.cell(row=1, column=col_idx).column_letter
        ws.column_dimensions[letter].width = min(max_width + 4, 50)

    output = io.BytesIO()
    wb.save(output)
    xlsx_bytes = output.getvalue()

    # Inject cell images if any
    if cell_images:
        from app.services.excel_image_extractor import inject_cell_images
        xlsx_bytes = inject_cell_images(xlsx_bytes, cell_images)

    output = io.BytesIO(xlsx_bytes)
    output.seek(0)
    return output


def _get_operator_name(session, operator_id: uuid.UUID) -> str:
    user = session.get(User, operator_id)
    return user.full_name or user.username if user else "未知"


def _get_export_group_name(session, export_group_id: uuid.UUID) -> str:
    group = session.get(ExportGroup, export_group_id)
    return group.name if group else "未知"


def _task_to_public(session, task) -> FilingTaskPublic:
    return FilingTaskPublic(
        id=task.id,
        task_name=task.task_name,
        qualification_count=task.qualification_count,
        port_count=task.port_count,
        export_group_name=_get_export_group_name(session, task.export_group_id),
        group_by_field=task.group_by_field,
        file_size=task.file_size,
        operator_name=_get_operator_name(session, task.operator_id),
        created_at=task.created_at,
    )


def _task_to_detail(session, task) -> FilingTaskDetail:
    public = _task_to_public(session, task)
    download_url = f"/api/v1/filing-tasks/{task.id}/download" if task.file_path else None

    return FilingTaskDetail(
        id=public.id,
        task_name=public.task_name,
        qualification_count=public.qualification_count,
        port_count=public.port_count,
        export_group_name=public.export_group_name,
        group_by_field=public.group_by_field,
        file_size=public.file_size,
        operator_name=public.operator_name,
        created_at=public.created_at,
        qualification_ids=task.qualification_ids,
        port_ids=task.port_ids,
        file_path=task.file_path,
        download_url=download_url,
    )


@router.get("/sub-port-availability")
def check_sub_port_availability(
    session: SessionDep,
    _current_user: CurrentUser,
    main_port_numbers: str = Query(...),
    range_start: int = Query(...),
    range_end: int = Query(...),
    mode: str = Query("random"),
) -> dict:
    from app.crud.filing_sub_port_usage import count_used_in_range
    result = {}
    for mpn in main_port_numbers.split(","):
        mpn = mpn.strip()
        if not mpn:
            continue
        if mode == "fixed_suffix":
            # 固定后缀模式不依赖范围：prefix 从 0 起递增，按 0-999999 统计已占用
            used = count_used_in_range(session, mpn, 0, 999999)
            result[mpn] = {
                "used": used,
                "total": 1_000_000,
                "available": max(0, 1_000_000 - used),
            }
        else:
            total = range_end - range_start + 1
            used = count_used_in_range(session, mpn, range_start, range_end)
            result[mpn] = {
                "used": used,
                "total": total,
                "available": max(0, total - used),
            }
    return result


@router.get("", response_model=FilingTasksPublic)
@router.get("/", include_in_schema=False, response_model=FilingTasksPublic)
def read_tasks(
    session: SessionDep,
    _current_user: CurrentUser,
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    start_date: date | None = None,
    end_date: date | None = None,
    keyword: str | None = None,
) -> Any:
    skip = (page - 1) * page_size
    items, total = list_filing_tasks(
        session=session, skip=skip, limit=page_size,
        start_date=start_date, end_date=end_date, keyword=keyword,
    )
    data = [_task_to_public(session, t) for t in items]
    return FilingTasksPublic(data=data, total=total, page=page, page_size=page_size)


@router.get("/{id}", response_model=FilingTaskDetail)
def read_task(*, session: SessionDep, id: uuid.UUID) -> Any:
    task = get_filing_task(session=session, id=id)
    if not task:
        raise HTTPException(status_code=404, detail="报备任务不存在")
    return _task_to_detail(session, task)


@router.post("", response_model=FilingTaskDetail)
@router.post("/", include_in_schema=False, response_model=FilingTaskDetail)
def create_task(*, session: SessionDep, create: FilingTaskCreate, current_user: CurrentUser, request: Request) -> Any:
    # 1. Validate export group exists
    export_group = get_export_group(session=session, id=create.export_group_id)
    if not export_group:
        raise HTTPException(status_code=404, detail="导出字段组不存在")

    # 2. Load qualifications by IDs
    if not create.qualification_ids:
        raise HTTPException(status_code=400, detail="至少选择一个资质")

    qual_id_strs = [str(qid) for qid in create.qualification_ids]

    qualifications = list(
        session.exec(
            select(QualificationInfo).where(QualificationInfo.id.in_(create.qualification_ids))  # type: ignore
        ).all()
    )
    if not qualifications:
        raise HTTPException(status_code=404, detail="未找到匹配的资质信息")

    # 3. Load ports by explicit IDs
    if not create.port_ids:
        raise HTTPException(status_code=400, detail="至少选择一个端口")

    selected_ports = list(
        session.exec(
            select(PortInfo).where(PortInfo.id.in_(create.port_ids))  # type: ignore
        ).all()
    )
    if len(selected_ports) != len(create.port_ids):
        raise HTTPException(status_code=400, detail="部分端口ID无效")
    selected_port_ids = [p.id for p in selected_ports]

    # 4. Create task record first (so we have an ID)
    task = crud_create_filing_task(
        session=session, create=create, operator_id=current_user.id
    )

    # 5. Load image attachments for selected qualifications
    # Map Chinese field_name → logical field name for export
    _CN_TO_LOGICAL_IMG = {
        "单位证件图片": "cert_image",
        "责任人身份证正面": "responsible_id_front",
        "责任人身份证反面": "responsible_id_back",
        "法人身份证正面": "handler_id_front",
        "法人身份证反面": "handler_id_back",
        "授权书图片": "auth_image",
        "签名举证附件": "signature_proof",
        "引流号码举证附件": "diversion_number_proof",
        "引流链接举证": "diversion_link_proof",
        "经办人现场照片": "handler_scene_photo",
    }
    qual_images: dict[uuid.UUID, dict[str, bytes]] = {}
    try:
        from app.crud.file_attachment import get_file_attachments_by_entity
        storage = get_storage()
        for q in qualifications:
            attachments = get_file_attachments_by_entity(
                session=session, entity_type="qualification_info", entity_id=q.id
            )
            img_map: dict[str, bytes] = {}
            for att in attachments:
                logical = _CN_TO_LOGICAL_IMG.get(att.field_name or "")
                if logical and att.stored_path:
                    try:
                        raw = storage.download(att.stored_path)
                        img_map[logical] = raw
                    except Exception:
                        pass
            if img_map:
                qual_images[q.id] = img_map
    except Exception:
        pass  # Image loading failure should not block export

    # 5.5 自动分配子端口号（可选）
    allocated_sub_ports: dict[tuple[uuid.UUID, str], str] = {}
    auto_allocate = create.auto_allocate_sub_ports

    if auto_allocate:
        try:
            try:
                mode = AllocationMode(create.allocation_mode or "random")
            except ValueError:
                raise HTTPException(status_code=400, detail=f"不支持的分配模式: {create.allocation_mode}")
            range_start = create.sub_port_range_start
            range_end = create.sub_port_range_end
            if mode == AllocationMode.fixed_suffix:
                # 固定后缀模式不使用范围参数
                range_start = range_start or 0
                range_end = range_end or 0
            else:
                if not (range_start and range_end):
                    raise HTTPException(status_code=400, detail="自动分配子端口时必须提供范围")
                if range_start > range_end:
                    raise HTTPException(status_code=400, detail="子端口范围起始必须 ≤ 结束")
                if mode == AllocationMode.random:
                    if range_end - range_start + 1 > MAX_RANGE_SIZE:
                        raise HTTPException(status_code=400, detail="子端口范围过大（最多 100 万个号码）")
                    # 固定 6 位格式（与默认值 100001 及规范一致），防止不同位数范围产生重复号码
                    if not (100000 <= range_start <= 999999 and 100000 <= range_end <= 999999):
                        raise HTTPException(status_code=400, detail="子端口范围必须是6位数字（100000-999999）")

            main_ports = [p for p in selected_ports if not p.sub_port_number]
            if not main_ports:
                raise HTTPException(status_code=400, detail="未找到主端口行（sub_port_number 为空的端口记录）")
            main_port_numbers = sorted({p.main_port_number for p in main_ports})

            allocation = allocate_sub_ports(
                session=session,
                main_port_numbers=main_port_numbers,
                range_start=range_start,
                range_end=range_end,
                qualifications=qualifications,
                operator_id=current_user.id,
                filing_task_id=task.id,
                mode=mode,
                fixed_suffix=create.fixed_suffix,
            )
            for mpn, pairs in allocation.items():
                for qual, num in pairs:
                    allocated_sub_ports[(qual.id, mpn)] = num
        except HTTPException:
            # 分配失败时清理已创建的报备任务记录
            delete_filing_task(session=session, db_obj=task)
            raise

    # 6. Generate Excel
    try:
        excel_bytes = generate_excel(
            qualifications=qualifications,
            ports=selected_ports,
            export_group=export_group,
            group_by_field=create.group_by_field,
            qual_images=qual_images,
            allocated_sub_ports=allocated_sub_ports,
            auto_allocate_sub_ports=auto_allocate,
        )
    except Exception as e:
        # 任务失败未产出文件：释放本任务已分配的子端口占用，避免永久烧号
        delete_usages_by_task(session=session, filing_task_id=task.id)
        delete_filing_task(session=session, db_obj=task)
        raise HTTPException(status_code=500, detail=f"生成Excel失败: {e}")

    file_data = excel_bytes.read()

    # 6. Upload to storage
    file_key = f"filing-exports/{task.id}.xlsx"
    try:
        storage = get_storage()
        storage.upload(key=file_key, data=file_data, content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    except Exception as e:
        # 任务失败未产出文件：释放本任务已分配的子端口占用，避免永久烧号
        delete_usages_by_task(session=session, filing_task_id=task.id)
        delete_filing_task(session=session, db_obj=task)
        raise HTTPException(status_code=500, detail=f"文件上传失败: {e}")

    # 7. Update task record with final values
    task.qualification_ids = qual_id_strs
    task.port_ids = [str(pid) for pid in selected_port_ids]
    task.port_count = len(selected_ports)
    task.qualification_count = len(qualifications)
    task.file_path = file_key
    task.file_size = len(file_data)
    session.add(task)
    session.commit()
    session.refresh(task)

    log_operation(session=session, user=current_user, user_ip=request.client.host if request.client else "", module="filing_tasks", action="create", target=task.task_name, detail=f"资质:{task.qualification_count} 端口:{task.port_count}")
    return _task_to_detail(session, task)


@router.delete("/{id}")
def delete_task(*, session: SessionDep, id: uuid.UUID, current_user: CurrentUser, request: Request) -> Message:
    task = get_filing_task(session=session, id=id)
    if not task:
        raise HTTPException(status_code=404, detail="报备任务不存在")

    # Delete MinIO file if exists
    if task.file_path:
        try:
            storage = get_storage()
            storage.delete(task.file_path)
        except Exception:
            pass  # File may already be gone

    target = task.task_name
    delete_filing_task(session=session, db_obj=task)
    log_operation(session=session, user=current_user, user_ip=request.client.host if request.client else "", module="filing_tasks", action="delete", target=target)
    return Message(message="报备任务删除成功")


@router.get("/{id}/download")
def download_task(*, session: SessionDep, id: uuid.UUID) -> Any:
    task = get_filing_task(session=session, id=id)
    if not task:
        raise HTTPException(status_code=404, detail="报备任务不存在")
    if not task.file_path:
        raise HTTPException(status_code=404, detail="文件不存在")

    try:
        storage = get_storage()
        content = storage.download(task.file_path)
        from urllib.parse import quote
        from pathlib import Path
        from fastapi.responses import Response

        ext = Path(task.file_path).suffix
        encoded_name = quote(f"{task.task_name}{ext}")
        return Response(
            content=content,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename*=UTF-8''{encoded_name}",
            },
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"文件下载失败: {e}")
