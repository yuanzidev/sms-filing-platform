"""Port info management routes."""

import io
import uuid
from typing import Any
from urllib.parse import quote

from fastapi import APIRouter, Depends, File, HTTPException, Query, Request, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from pydantic import BaseModel

from app.api.deps import CurrentUser, SessionDep, get_current_active_superuser
from app.crud.port_info import (
    create_port_info,
    delete_port_info,
    get_port_info,
    list_port_infos,
    update_port_info,
)
from app.models import (
    Message,
    PortInfo,
    PortInfoCreate,
    PortInfoPublic,
    PortInfosPublic,
    PortInfoUpdate,
)
from app.services.excel_image_extractor import (
    extract_cell_images_from_xlsx,
    extract_images_from_xlsx,
    upload_import_images,
)
from app.services.operation_log import log_operation

router = APIRouter(
    prefix="/port-info",
    tags=["port-info"],
    dependencies=[Depends(get_current_active_superuser)],
)

_PORT_HEADER_TO_FIELD = {
    "运营商": "carrier",
    "主端口号": "main_port_number",
    "主端口备案公司": "enterprise_name",
    "子端口号": "sub_port_number",
    "码号使用范围": "port_range",
    "接入省": "province",
    "接入地市": "city",
    "端口类型": "port_type",
    "操作类型": "operation_type",
    "端口入网时间": "port_activation_date",
    "是否允许自行扩展": "allow_self_extension",
    "运营商接入机房及设备": "carrier_room",
    "企业接入机房及设备": "enterprise_room",
    "是否具有授权书": "has_authorization",
    "授权书": "authorization_letter",
    "授权开始日期": "auth_start_date",
    "授权结束日期": "auth_end_date",
    "集团编码": "group_code",
    "所属地区": "region",
    "其他接入机房说明": "other_room_description",
    "是否绿色通道": "is_green_channel",
    "黑白名单类型": "blacklist_whitelist_type",
    "端口审核表": "audit_form",
    "客户类型": "customer_type",
    "基础电信企业ID": "basic_telecom_enterprise_id",
}

_PORT_HEADERS = [
    "运营商",
    "主端口号",
    "主端口备案公司",
    "子端口号",
    "码号使用范围",
    "接入省",
    "接入地市",
    "端口类型",
    "操作类型",
    "端口入网时间",
    "是否允许自行扩展",
    "运营商接入机房及设备",
    "企业接入机房及设备",
    "是否具有授权书",
    "授权书",
    "授权开始日期",
    "授权结束日期",
    "集团编码",
    "所属地区",
    "其他接入机房说明",
    "是否绿色通道",
    "黑白名单类型",
    "端口审核表",
    "客户类型",
    "基础电信企业ID",
    "授权书图片",
]


@router.get("/template")
def download_port_info_template() -> Any:
    from openpyxl.styles import Font
    from PIL import Image, ImageDraw
    from app.services.excel_image_extractor import inject_cell_images

    wb = Workbook()
    ws = wb.active
    ws.title = "端口信息导入模板"

    for col_idx, header in enumerate(_PORT_HEADERS, 1):
        ws.cell(row=1, column=col_idx, value=header)

    example_data = [
        "中国移动",
        "10690001",
        "示例企业有限公司",
        "0001",
        "全国",
        "广东",
        "深圳",
        "短信",
        "新增",
        "2024-01-15",
        "是",
        "运营商XX机房-XX设备",
        "企业XX机房-XX设备",
        "是",
        "授权书编号001",
        "2024-01-01",
        "2025-12-31",
        "G001",
        "华南地区",
        "备用机房A",
        "否",
        "黑名单",
        "已审核",
        "企业客户",
    ]
    for col_idx, val in enumerate(example_data, 1):
        ws.cell(row=2, column=col_idx, value=val)

    instructions = wb.create_sheet("填写说明")
    instructions.cell(row=1, column=1, value="Excel 导入图片填写说明").font = Font(
        bold=True, size=14
    )
    notes = [
        "1. 请勿修改表头行（第一行）的列标题",
        "2. 每条数据填写一行，从第二行开始",
        "3. 图片列（授权书图片）用于存放授权书扫描件等图片文件",
        "4. 插入方法：右键单元格 ->「插入图片」->「放置在单元格中」-> 选择图片文件",
        "5. 也可将图片直接拖入到图片列的单元格中",
        "6. 系统会自动提取每行单元格内嵌的图片，并与对应字段关联",
        "7. 支持的图片格式：PNG、JPEG、GIF、BMP、WEBP，单张不超过 10MB",
        "8. 操作类型、集团编码：选填",
        "9. 授权书图片列支持插入图片文件；导出时图片会嵌入 Excel 单元格",
    ]
    for i, note in enumerate(notes, 2):
        instructions.cell(row=i, column=1, value=note)

    output = io.BytesIO()
    wb.save(output)
    xlsx_bytes = output.getvalue()

    sample_img = Image.new("RGB", (120, 60), color=(220, 230, 241))
    draw = ImageDraw.Draw(sample_img)
    draw.text((10, 20), "示例\n请替换", fill=(50, 50, 50))
    img_buf = io.BytesIO()
    sample_img.save(img_buf, format="PNG")

    # Authorization image column = column 26 (1-based) = "Z2"
    cell_images = {"Z2": img_buf.getvalue()}
    xlsx_bytes = inject_cell_images(xlsx_bytes, cell_images)

    return StreamingResponse(
        io.BytesIO(xlsx_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename*=UTF-8''{quote('端口信息导入模板_v2.xlsx')}"
        },
    )


@router.post("/import/preview")
def preview_port_info_import(file: UploadFile = File(...)) -> Any:
    """解析 Excel 前 5 行数据并返回预览，供导入前核对表头与数据。"""
    if not file.filename or not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="仅支持 .xlsx 或 .xls 文件")
    content = file.file.read()
    try:
        wb = load_workbook(io.BytesIO(content))
    except Exception:
        raise HTTPException(status_code=400, detail="无法解析 Excel 文件")
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 1:
        raise HTTPException(status_code=400, detail="文件为空")

    header_to_field = _PORT_HEADER_TO_FIELD
    header_row = [str(c) if c else "" for c in rows[0]]
    col_map: dict[str, int] = {}
    for col_idx, h in enumerate(header_row):
        if h in header_to_field:
            col_map[header_to_field[h]] = col_idx

    unrecognized = [
        h
        for h in header_row
        if h and h not in header_to_field and h not in ("", "None")
    ]

    preview_rows = []
    for row in rows[1:6]:  # first 5 data rows
        if all(c is None or str(c).strip() == "" for c in row):
            continue
        row_data = {}
        for field_name, col_idx in col_map.items():
            v = row[col_idx] if col_idx < len(row) else None
            row_data[field_name] = (
                str(v).strip() if v is not None and str(v).strip() else None
            )
        preview_rows.append(row_data)

    return {
        "headers": header_row,
        "rows": preview_rows,
        "unrecognized_headers": unrecognized,
        "total_data_rows": len(
            [
                r
                for r in rows[1:]
                if not all(c is None or str(c).strip() == "" for c in r)
            ]
        ),
    }


@router.post("/import")
def import_port_infos(*, session: SessionDep, file: UploadFile = File(...)) -> Any:
    if not file.filename or not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="仅支持 .xlsx 或 .xls 文件")

    content = file.file.read()
    try:
        wb = load_workbook(io.BytesIO(content))
    except Exception:
        raise HTTPException(
            status_code=400, detail="无法解析 Excel 文件，请检查文件格式"
        )

    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 1:
        raise HTTPException(status_code=400, detail="文件为空，请导入有效的 Excel 文件")

    header_to_field = _PORT_HEADER_TO_FIELD

    header_row = [str(c) if c else "" for c in rows[0]]
    col_map: dict[str, int] = {}
    for col_idx, h in enumerate(header_row):
        if h in header_to_field:
            col_map[header_to_field[h]] = col_idx

    unrecognized_headers = [
        h
        for h in header_row
        if h and h not in header_to_field and h not in ("", "None")
    ]

    required_fields = ["carrier", "main_port_number", "enterprise_name", "port_type"]
    missing = [
        h
        for h, f in header_to_field.items()
        if f in required_fields and f not in col_map
    ]
    if missing:
        raise HTTPException(
            status_code=400,
            detail=f"模板不匹配：缺少必填列「{'」「'.join(missing)}」。请确认使用了正确的端口信息导入模板（首页 → 端口管理 → 下载模板）",
        )

    # Phase 1: Validate all rows, collect errors
    objects: list[PortInfo] = []
    errors: list[dict] = []
    data_row_indices: list[int] = []

    for row_idx, row in enumerate(rows[1:], start=2):
        if all(c is None or str(c).strip() == "" for c in row):
            continue

        def cell(col_name: str) -> str | None:
            if col_name not in col_map:
                return None
            idx = col_map[col_name]
            if idx >= len(row):
                return None
            v = row[idx]
            if v is None or str(v).strip() == "":
                return None
            return str(v).strip()

        def parse_bool(col_name: str) -> bool | None:
            v = cell(col_name)
            if v is None:
                return None
            return v in ("是", "true", "True", "1", "TRUE")

        row_errors: list[dict] = []

        def parse_date(col_name: str, cn_name: str):
            if col_name not in col_map:
                return None
            idx = col_map[col_name]
            if idx >= len(row):
                return None
            v = row[idx]
            if v is None:
                return None
            from datetime import date, datetime

            if isinstance(v, datetime):
                return v.date()
            if isinstance(v, date):
                return v
            s = str(v).strip()
            if not s:
                return None
            try:
                return date.fromisoformat(s)
            except ValueError:
                row_errors.append(
                    {
                        "row": row_idx,
                        "field": cn_name,
                        "value": s,
                        "reason": "日期格式无效",
                        "suggestion": "请使用 YYYY-MM-DD 格式",
                    }
                )
                return None

        carrier = cell("carrier")
        if not carrier:
            row_errors.append(
                {
                    "row": row_idx,
                    "field": "运营商",
                    "value": "",
                    "reason": "运营商不能为空",
                    "suggestion": "请填写运营商",
                }
            )

        main_port_number = cell("main_port_number")
        if not main_port_number:
            row_errors.append(
                {
                    "row": row_idx,
                    "field": "主端口号",
                    "value": "",
                    "reason": "主端口号不能为空",
                    "suggestion": "请填写主端口号",
                }
            )

        enterprise_name = cell("enterprise_name")
        if not enterprise_name:
            row_errors.append(
                {
                    "row": row_idx,
                    "field": "主端口备案公司",
                    "value": "",
                    "reason": "企业名称不能为空",
                    "suggestion": "请填写主端口备案公司",
                }
            )

        port_type = cell("port_type")
        if not port_type:
            row_errors.append(
                {
                    "row": row_idx,
                    "field": "端口类型",
                    "value": "",
                    "reason": "端口类型不能为空",
                    "suggestion": "请填写端口类型",
                }
            )

        # Validate booleans
        for bool_field, cn_name in [
            ("allow_self_extension", "是否允许自行扩展"),
            ("has_authorization", "是否具有授权书"),
            ("is_green_channel", "是否绿色通道"),
        ]:
            v = cell(bool_field)
            if v and v not in (
                "是",
                "否",
                "true",
                "True",
                "1",
                "TRUE",
                "false",
                "False",
                "0",
                "FALSE",
            ):
                row_errors.append(
                    {
                        "row": row_idx,
                        "field": cn_name,
                        "value": v,
                        "reason": "布尔字段值无效",
                        "suggestion": "请填写「是」或「否」",
                    }
                )

        # Validate dates (collects errors into row_errors)
        parse_date("port_activation_date", "端口入网时间")
        parse_date("auth_start_date", "授权开始日期")
        parse_date("auth_end_date", "授权结束日期")

        if row_errors:
            errors.extend(row_errors)
        else:
            objects.append(
                PortInfo(
                    carrier=carrier,
                    main_port_number=main_port_number,
                    enterprise_name=enterprise_name,
                    sub_port_number=cell("sub_port_number"),
                    port_range=cell("port_range"),
                    province=cell("province"),
                    city=cell("city"),
                    port_type=port_type,
                    operation_type=cell("operation_type"),
                    port_activation_date=parse_date(
                        "port_activation_date", "端口入网时间"
                    ),
                    allow_self_extension=parse_bool("allow_self_extension"),
                    carrier_room=cell("carrier_room"),
                    enterprise_room=cell("enterprise_room"),
                    has_authorization=parse_bool("has_authorization"),
                    authorization_letter=cell("authorization_letter"),
                    auth_start_date=parse_date("auth_start_date", "授权开始日期"),
                    auth_end_date=parse_date("auth_end_date", "授权结束日期"),
                    group_code=cell("group_code"),
                    region=cell("region"),
                    other_room_description=cell("other_room_description"),
                    is_green_channel=parse_bool("is_green_channel"),
                    blacklist_whitelist_type=cell("blacklist_whitelist_type"),
                    audit_form=cell("audit_form"),
                    customer_type=cell("customer_type"),
                    basic_telecom_enterprise_id=cell("basic_telecom_enterprise_id"),
                )
            )
            data_row_indices.append(row_idx)

    # Phase 2: If no valid rows, return all errors
    if not objects and errors:
        return {
            "total": len(objects) + len({e["row"] for e in errors}),
            "success_count": 0,
            "error_count": len(errors),
            "errors": errors,
            "unrecognized_headers": unrecognized_headers,
        }

    if not objects:
        raise HTTPException(status_code=400, detail="文件中没有有效数据")

    # Phase 3: Write valid rows + extract images with fixed indices
    session.add_all(objects)
    session.flush()

    warnings: list[str] = []
    if file.filename.endswith(".xlsx"):
        all_images: list = []
        try:
            all_images.extend(
                extract_cell_images_from_xlsx(
                    content,
                    headers=header_row,
                    data_row_indices=data_row_indices,
                )
            )
        except Exception as e:
            warnings.append(f"单元格图片提取失败: {e}")
        try:
            all_images.extend(
                extract_images_from_xlsx(
                    content,
                    headers=header_row,
                    data_row_indices=data_row_indices,
                )
            )
        except Exception as e:
            warnings.append(f"浮动图片提取失败: {e}")
        if all_images:
            _, img_warnings, img_errors = upload_import_images(
                images=all_images,
                objects=objects,
                entity_type="port_info",
                session=session,
            )
            warnings.extend(img_warnings)
            errors.extend(img_errors)

    session.commit()

    return {
        "total": len(objects) + len({e["row"] for e in errors}),
        "success_count": len(objects),
        "error_count": len(errors),
        "errors": errors,
        "warnings": warnings,
        "unrecognized_headers": unrecognized_headers,
    }


class ImportErrorReport(BaseModel):
    errors: list[dict]


@router.post("/import/error-report")
def download_import_error_report(body: ImportErrorReport) -> Any:
    """Generate an Excel file highlighting import errors."""
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "导入错误报告"

    headers = ["行号", "字段", "原值", "失败原因", "修复建议"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)

    red_fill = PatternFill(start_color="FFD7D7", end_color="FFD7D7", fill_type="solid")
    for i, err in enumerate(body.errors, 2):
        ws.cell(row=i, column=1, value=err.get("row"))
        ws.cell(row=i, column=2, value=err.get("field"))
        ws.cell(row=i, column=3, value=err.get("value"))
        ws.cell(row=i, column=4, value=err.get("reason"))
        ws.cell(row=i, column=5, value=err.get("suggestion"))
        for col in range(1, 6):
            ws.cell(row=i, column=col).fill = red_fill

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename*=UTF-8''{quote('导入错误报告.xlsx')}"
        },
    )


@router.get("", response_model=PortInfosPublic)
@router.get("/", include_in_schema=False, response_model=PortInfosPublic)
def read_port_infos(
    session: SessionDep,
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=500),
    carrier: str | None = None,
    province: str | None = None,
    keyword: str | None = None,
    city: str | None = None,
    port_type: str | None = None,
    main_port_number: str | None = None,
) -> Any:
    skip = (page - 1) * page_size
    items, total = list_port_infos(
        session=session,
        skip=skip,
        limit=page_size,
        carrier=carrier,
        province=province,
        keyword=keyword,
        city=city,
        port_type=port_type,
        main_port_number=main_port_number,
    )
    return PortInfosPublic(data=items, total=total, page=page, page_size=page_size)


@router.post("", response_model=PortInfoPublic)
@router.post("/", include_in_schema=False, response_model=PortInfoPublic)
def create_port_info_endpoint(
    *,
    session: SessionDep,
    create: PortInfoCreate,
    current_user: CurrentUser,
    request: Request,
) -> Any:
    try:
        result = create_port_info(session=session, create=create)
    except Exception as e:
        error_str = str(e)
        if "unique" in error_str.lower() or "duplicate" in error_str.lower():
            raise HTTPException(
                status_code=409,
                detail={
                    "field": "main_port_number",
                    "reason": "主端口号已存在",
                    "suggestion": "请使用不同的主端口号，或先查询已有端口信息",
                },
            )
        raise HTTPException(
            status_code=500,
            detail={"field": "", "reason": error_str, "suggestion": "请联系管理员"},
        )
    log_operation(
        session=session,
        user=current_user,
        user_ip=request.client.host if request.client else "",
        module="port_info",
        action="create",
        target=f"{result.main_port_number or result.sub_port_number or result.id}",
    )
    return result


@router.get("/{id}", response_model=PortInfoPublic)
def read_port_info(*, session: SessionDep, id: uuid.UUID) -> Any:
    db_obj = get_port_info(session=session, id=id)
    if not db_obj:
        raise HTTPException(status_code=404, detail="端口信息不存在")
    return db_obj


@router.patch("/{id}", response_model=PortInfoPublic)
def update_port_info_endpoint(
    *,
    session: SessionDep,
    id: uuid.UUID,
    update: PortInfoUpdate,
    current_user: CurrentUser,
    request: Request,
) -> Any:
    db_obj = get_port_info(session=session, id=id)
    if not db_obj:
        raise HTTPException(status_code=404, detail="端口信息不存在")
    result = update_port_info(session=session, db_obj=db_obj, update=update)
    log_operation(
        session=session,
        user=current_user,
        user_ip=request.client.host if request.client else "",
        module="port_info",
        action="update",
        target=f"{result.main_port_number or result.sub_port_number or id}",
    )
    return result


@router.delete("/{id}")
def delete_port_info_endpoint(
    *, session: SessionDep, id: uuid.UUID, current_user: CurrentUser, request: Request
) -> Message:
    db_obj = get_port_info(session=session, id=id)
    if not db_obj:
        raise HTTPException(status_code=404, detail="端口信息不存在")
    target = f"{db_obj.main_port_number or db_obj.sub_port_number or id}"
    delete_port_info(session=session, db_obj=db_obj)
    log_operation(
        session=session,
        user=current_user,
        user_ip=request.client.host if request.client else "",
        module="port_info",
        action="delete",
        target=target,
    )
    return Message(message="端口信息删除成功")
