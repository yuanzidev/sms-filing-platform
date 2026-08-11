"""Export group API routes."""

import io
import uuid
from dataclasses import asdict
from pathlib import PurePath
from typing import Any
from urllib.parse import quote

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook

from app.api.deps import SessionDep, get_current_active_superuser
from app.crud.export_group import (
    create_export_group,
    delete_export_group,
    get_export_group,
    list_export_groups,
    update_export_group,
)
from app.models import (
    ExportGroup,
    ExportGroupCreate,
    ExportGroupField,
    ExportGroupPublic,
    ExportGroupsPublic,
    ExportGroupUpdate,
    Message,
)
from app.services.export_field_registry import all_fields

router = APIRouter(
    prefix="/export-groups",
    tags=["export-groups"],
    dependencies=[Depends(get_current_active_superuser)],
)


def _cell_text(value: Any) -> str:
    return str(value).strip() if value is not None else ""


def _filename_stem(filename: str | None) -> str:
    if not filename:
        return "导入字段组"
    return PurePath(filename).stem.strip() or "导入字段组"


def _find_header_index(headers: list[str], candidates: set[str]) -> int | None:
    for idx, header in enumerate(headers):
        if header in candidates:
            return idx
    return None


@router.get("/registry", response_model=list[dict])
def read_field_registry() -> Any:
    return [{**asdict(f), "id": f.name} for f in all_fields()]


@router.get("/registry/template")
def download_registry_template() -> Any:
    wb = Workbook()
    ws = wb.active
    ws.title = "字段编码对照表"
    ws.cell(row=1, column=1, value="字段编码")
    ws.cell(row=1, column=2, value="字段名称")
    ws.cell(row=1, column=3, value="所属分组")
    for i, entry in enumerate(all_fields(), 2):
        ws.cell(row=i, column=1, value=entry.name)
        ws.cell(row=i, column=2, value=entry.label)
        ws.cell(row=i, column=3, value=entry.group)
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename*=UTF-8''{quote('字段编码对照表.xlsx')}"
        },
    )


@router.get("", response_model=ExportGroupsPublic)
@router.get("/", include_in_schema=False, response_model=ExportGroupsPublic)
def read_export_groups(session: SessionDep) -> Any:
    groups = list_export_groups(session=session)
    return ExportGroupsPublic(data=groups, count=len(groups))


@router.post("", response_model=ExportGroupPublic)
@router.post("/", include_in_schema=False, response_model=ExportGroupPublic)
def create_export_group_endpoint(
    *, session: SessionDep, create: ExportGroupCreate
) -> Any:
    db_obj = create_export_group(session=session, create=create)
    return db_obj


@router.get("/{id}", response_model=ExportGroupPublic)
def read_export_group(*, session: SessionDep, id: uuid.UUID) -> Any:
    db_obj = get_export_group(session=session, id=id)
    if not db_obj:
        raise HTTPException(status_code=404, detail="导出分组不存在")
    return db_obj


@router.patch("/{id}", response_model=ExportGroupPublic)
def update_export_group_endpoint(
    *, session: SessionDep, id: uuid.UUID, update: ExportGroupUpdate
) -> Any:
    db_obj = get_export_group(session=session, id=id)
    if not db_obj:
        raise HTTPException(status_code=404, detail="导出分组不存在")
    db_obj = update_export_group(session=session, db_obj=db_obj, update=update)
    return db_obj


@router.delete("/{id}")
def delete_export_group_endpoint(*, session: SessionDep, id: uuid.UUID) -> Message:
    db_obj = get_export_group(session=session, id=id)
    if not db_obj:
        raise HTTPException(status_code=404, detail="导出分组不存在")
    delete_export_group(session=session, db_obj=db_obj)
    return Message(message="导出分组删除成功")


@router.get("/{id}/export")
def export_export_group(*, session: SessionDep, id: uuid.UUID) -> Any:
    group = get_export_group(session=session, id=id)
    if not group:
        raise HTTPException(status_code=404, detail="字段组不存在")

    wb = Workbook()
    ws = wb.active
    ws.title = "字段组"

    headers = ["字段组名称", "字段编码", "字段名称", "字段顺序", "是否启用"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    for i, field in enumerate(sorted(group.fields, key=lambda f: f.sort_order)):
        ws.cell(row=i + 2, column=1, value=group.name)
        ws.cell(row=i + 2, column=2, value=field.field_name)
        ws.cell(row=i + 2, column=3, value=field.field_label)
        ws.cell(row=i + 2, column=4, value=field.sort_order)
        ws.cell(row=i + 2, column=5, value="是")

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename*=UTF-8''{quote(f'{group.name}.xlsx')}"
        },
    )


@router.post("/import")
def import_export_group(*, session: SessionDep, file: UploadFile = File(...)) -> Any:
    if not file.filename or not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="仅支持 .xlsx 或 .xls 文件")
    content = file.file.read()
    try:
        wb = load_workbook(io.BytesIO(content))
    except Exception:
        raise HTTPException(status_code=400, detail="无法解析 Excel 文件")
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail="文件中没有数据")

    registry_map = {f.name: f for f in all_fields()}
    errors: list[dict] = []

    header = [_cell_text(value) for value in rows[0]]
    group_col = _find_header_index(header, {"字段组名称", "分组名称", "字段组"})
    code_col = _find_header_index(
        header, {"字段编码", "字段代码", "字段code", "field_code", "field_name"}
    )
    label_col = _find_header_index(header, {"字段名称", "字段标签", "field_label"})
    order_col = _find_header_index(header, {"字段顺序", "排序", "sort_order"})

    if group_col is not None or code_col is not None:
        data_rows = rows[1:]
        code_col = code_col if code_col is not None else 1
    elif len(rows[0]) == 1:
        data_rows = rows
        code_col = 0
    elif len(rows[0]) > 1 and _cell_text(rows[0][1]) in registry_map:
        data_rows = rows
        group_col = 0
        code_col = 1
        label_col = 2
        order_col = 3
    else:
        data_rows = rows[1:]
        group_col = 0
        code_col = 1
        label_col = 2
        order_col = 3

    group_name = ""
    if group_col is not None:
        group_name = next(
            (
                _cell_text(row[group_col])
                for row in data_rows
                if len(row) > group_col and _cell_text(row[group_col])
            ),
            "",
        )
    if not group_name:
        group_name = _filename_stem(file.filename)
    if not group_name:
        raise HTTPException(status_code=400, detail="字段组名称不能为空")

    fields = []
    row_offset = 2 if data_rows is not rows else 1
    for row_idx, row in enumerate(data_rows, start=row_offset):
        field_code = _cell_text(row[code_col]) if len(row) > code_col else ""
        if not field_code:
            continue
        if field_code not in registry_map:
            errors.append(
                {
                    "row": row_idx,
                    "field": "字段编码",
                    "value": field_code,
                    "reason": "字段编码不存在于注册表",
                    "suggestion": "请下载字段编码对照表核对",
                }
            )
            continue
        label = (
            _cell_text(row[label_col])
            if label_col is not None
            and len(row) > label_col
            and _cell_text(row[label_col])
            else registry_map[field_code].label
        )
        order = (
            int(row[order_col])
            if order_col is not None and len(row) > order_col and row[order_col]
            else len(fields)
        )
        fields.append(
            {"field_name": field_code, "field_label": label, "sort_order": order}
        )

    if errors:
        return {"success_count": 0, "error_count": len(errors), "errors": errors}

    if not fields:
        raise HTTPException(status_code=400, detail="文件中没有有效字段数据")

    group = ExportGroup(name=group_name)
    for f in fields:
        group.fields.append(
            ExportGroupField(
                field_name=f["field_name"],
                field_label=f["field_label"],
                sort_order=f["sort_order"],
            )
        )
    session.add(group)
    session.commit()
    session.refresh(group)
    return {"success_count": 1, "group_name": group.name, "field_count": len(fields)}
